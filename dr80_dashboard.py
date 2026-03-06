"""
DR80 Tracking Dashboard
KTB Securities — Depositary Receipt Operations
"""

import streamlit as st
import pandas as pd
import numpy as np
import plotly.graph_objects as go
import io, os, warnings, json, re, base64, requests
from datetime import datetime, timedelta
import yfinance as yf
import psycopg2
from psycopg2.extras import RealDictCursor

warnings.filterwarnings("ignore")

st.set_page_config(
    page_title="DR80 Dashboard",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── CSS ────────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=IBM+Plex+Mono:wght@400;600&family=IBM+Plex+Sans:wght@300;400;600&display=swap');
html, body, [class*="css"] { font-family: 'IBM Plex Sans', sans-serif; }
.stApp { background: #0a0e1a; color: #e2e8f0; }
section[data-testid="stSidebar"] { background: #0d1221 !important; border-right: 1px solid #1e2d4a; }
section[data-testid="stSidebar"] * { color: #94a3b8 !important; }
.metric-card { background: linear-gradient(135deg,#111827,#0f1729); border:1px solid #1e2d4a; border-radius:8px; padding:16px 20px; margin-bottom:8px; position:relative; overflow:hidden; }
.metric-card::before { content:''; position:absolute; top:0; left:0; width:3px; height:100%; background:#3b82f6; }
.metric-card.green::before { background:#10b981; }
.metric-card.red::before { background:#ef4444; }
.metric-label { font-family:'IBM Plex Mono',monospace; font-size:0.75rem; text-transform:uppercase; letter-spacing:0.1em; color:#475569; margin-bottom:4px; }
.metric-value { font-family:'IBM Plex Mono',monospace; font-size:1.8rem; font-weight:600; color:#e2e8f0; }
.metric-sub { font-size:0.85rem; color:#64748b; margin-top:2px; }
.section-header { font-family:'IBM Plex Mono',monospace; font-size:0.7rem; text-transform:uppercase; letter-spacing:0.12em; color:#3b82f6; border-bottom:1px solid #1e2d4a; padding-bottom:8px; margin-bottom:16px; margin-top:24px; }
.dashboard-title { font-family:'IBM Plex Mono',monospace; font-size:1.4rem; font-weight:600; color:#e2e8f0; letter-spacing:-0.02em; }
.dashboard-sub { font-family:'IBM Plex Mono',monospace; font-size:0.7rem; color:#334155; letter-spacing:0.05em; }
.stMultiSelect span[data-baseweb="tag"] { background:rgba(59,130,246,0.2)!important; border:1px solid rgba(59,130,246,0.4)!important; color:#93c5fd!important; }
hr { border-color:#1e2d4a; }
.stButton > button { background:#1e2d4a; color:#94a3b8; border:1px solid #2d3f5e; border-radius:6px; font-family:'IBM Plex Mono',monospace; font-size:0.75rem; }
.stButton > button:hover { background:#2d3f5e; color:#e2e8f0; border-color:#3b82f6; }
</style>
""", unsafe_allow_html=True)

# ── Constants ──────────────────────────────────────────────────────────────────
PERIODS = ["YTD", "1M", "3M", "6M", "1Y", "3Y", "5Y"]
SECTORS = [
    "Semiconductor/ AI", "Techonology", "Precious Metal",
    "Energy", "Consumer discretionary", "Consumer defensive",
    "Defense", "ETF"
]
DEFAULT_FILE = "DR80_Tracking.xlsx"


# ── Ticker conversion ──────────────────────────────────────────────────────────
def bbg_to_yahoo(bbg: str):
    t = bbg.strip()
    parts = t.rsplit(" ", 2)
    if len(parts) < 3 or parts[2].upper() != "EQUITY":
        return t or None
    code, exch = parts[0].strip(), parts[1].strip().upper()
    if exch == "TB":
        return None  # Thai DRs not on Yahoo
    if exch == "US":
        return code
    if exch == "HK":
        try:
            return f"{int(code):04d}.HK"
        except ValueError:
            return f"{code}.HK"
    if exch == "JP":
        return f"{code}.T"
    if exch == "CH":
        try:
            int(code)
            return f"{code}.SS" if code.startswith("6") else f"{code}.SZ"
        except ValueError:
            return f"{code}.SS"
    mapping = {"SP": ".SI", "LN": ".L", "FP": ".PA", "GR": ".DE", "AU": ".AX"}
    suffix = mapping.get(exch, "")
    return f"{code}{suffix}" if suffix else code


def short_ticker(bbg: str) -> str:
    """Strip exchange suffix, return raw code."""
    for suffix in [" TB Equity", " US Equity", " HK Equity", " JP Equity",
                   " CH Equity", " SP Equity", " LN Equity", " FP Equity",
                   " GR Equity", " AU Equity"]:
        bbg = bbg.replace(suffix, "")
    return bbg.strip()


@st.cache_data(show_spinner=False)
def display_label(bbg: str, name: str, max_len: int = 15) -> str:
    """
    For numeric-code tickers (e.g. 300476, 9984), return a shortened company name.
    For alpha tickers (AAPL, NVDA80), return the stripped ticker code.
    Always returns a plain STRING — never a number — so Plotly treats it as a category.
    """
    code = short_ticker(bbg)
    # If the code is purely numeric, use shortened company name instead
    try:
        int(code)
        # It's a number — use company name, trimmed
        label = name.strip()
        if len(label) > max_len:
            label = label[:max_len].rstrip() + "…"
        return label
    except ValueError:
        # Alpha ticker — strip the "80" suffix for DR80 tickers to keep it short
        return code


# ── Excel parsing ──────────────────────────────────────────────────────────────
@st.cache_data(show_spinner=False)
def _parse_sheet(df_raw: pd.DataFrame) -> pd.DataFrame:
    """Parse a DR80-structured sheet (sector headers + ticker rows) into a DataFrame."""
    records = []
    current_sector = "Unknown"
    for _, row in df_raw.iterrows():
        c0 = row[0]
        c1 = str(row[1]) if pd.notna(row[1]) else ""
        c2 = str(row[2]) if pd.notna(row[2]) else ""
        c3 = str(row[3]) if pd.notna(row[3]) else ""
        if pd.isna(c0) and c2 == "name" and c1 not in ["", "nan"]:
            current_sector = c1.strip()
            continue
        if c1 in ["", "nan"] or c2 in ["", "nan", "id", "name"] or c1.startswith("Unnamed"):
            continue
        perf = {}
        for j, p in enumerate(PERIODS):
            v = row[4 + j] if (4 + j) < len(row) else None
            perf[p] = float(v) if pd.notna(v) else None
        records.append({
            "BBG_Ticker": c1.strip(),
            "Yahoo_Ticker": bbg_to_yahoo(c1.strip()),
            "Name": c2.strip(),
            "Sector": current_sector,
            "Quarter": c3.strip() if c3 not in ["nan", ""] else None,
            "Is_DR80": c1.strip().endswith("80 TB Equity"),
            **perf,
        })
    return pd.DataFrame(records)


@st.cache_data(show_spinner=False)
def parse_excel(file_bytes: bytes) -> pd.DataFrame:
    df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Current DR80", header=None)
    return _parse_sheet(df_raw)


@st.cache_data(show_spinner=False)
def parse_competitors(file_bytes: bytes) -> pd.DataFrame:
    """Parse the 'Competitors' sheet if it exists, else return empty DataFrame."""
    try:
        xl = pd.ExcelFile(io.BytesIO(file_bytes))
        if "Competitors" not in xl.sheet_names:
            return pd.DataFrame()
        df_raw = pd.read_excel(io.BytesIO(file_bytes), sheet_name="Competitors", header=None)
        df = _parse_sheet(df_raw)
        df["Is_DR80"] = False
        return df
    except Exception:
        return pd.DataFrame()


def match_competitor_to_dr80(comp_bbg: str, comp_sector: str, dr80_df: pd.DataFrame) -> str | None:
    """
    Match a competitor to a DR80 security.
    1. Try exact base-code match (e.g. 'NVDA' in comp matches 'NVDA80 TB Equity' in DR80)
    2. Fall back to sector match — return sector label
    """
    comp_code = short_ticker(comp_bbg).upper()
    # Remove trailing digits/suffix variants
    for row in dr80_df[dr80_df["Is_DR80"]].itertuples():
        dr_code = short_ticker(row.BBG_Ticker).upper()
        # Strip "80" from DR80 ticker for comparison
        dr_base = dr_code.replace("80", "")
        if comp_code == dr_base or comp_code == dr_code:
            return row.BBG_Ticker
    # No exact match — return sector as the group key
    return comp_sector


# ── Yahoo Finance fetch ────────────────────────────────────────────────────────
def fetch_returns_yahoo(tickers: list) -> dict:
    try:
        import yfinance as yf
    except ImportError:
        st.error("yfinance not installed. Run: pip install yfinance")
        return {}

    today = datetime.today()
    start = (today - timedelta(days=365 * 5 + 30)).strftime("%Y-%m-%d")
    valid = [t for t in tickers if t]
    if not valid:
        return {}

    _status = st.empty()
    _status.caption("⏳ Connecting to Yahoo Finance...")
    try:
        raw = yf.download(valid, start=start, end=today.strftime("%Y-%m-%d"),
                          auto_adjust=True, progress=False)
        prices = raw["Close"] if "Close" in raw.columns else raw
        _status.empty()
    except Exception as e:
        st.error(f"Download failed: {e}")
        _status.empty()
        return {}

    def pct_chg(series, from_dt):
        s = series.dropna()
        idx = s.index.searchsorted(pd.Timestamp(from_dt))
        if idx >= len(s):
            return None
        sp, ep = s.iloc[idx], s.iloc[-1]
        return (ep / sp - 1) * 100 if sp != 0 else None

    period_offsets = {
        "YTD": datetime(today.year, 1, 1),
        "1M":  today - timedelta(days=30),
        "3M":  today - timedelta(days=91),
        "6M":  today - timedelta(days=182),
        "1Y":  today - timedelta(days=365),
        "3Y":  today - timedelta(days=365 * 3),
        "5Y":  today - timedelta(days=365 * 5),
    }

    results = {}
    for i, ticker in enumerate(valid):
        _status.caption(f"⏳ Processing {ticker} ({i+1}/{len(valid)})…")
        try:
            s = prices[ticker] if (isinstance(prices, pd.DataFrame) and ticker in prices.columns) else prices
            results[ticker] = {p: pct_chg(s, dt) for p, dt in period_offsets.items()}
        except Exception:
            results[ticker] = {p: None for p in PERIODS}

    _status.empty()
    return results


def fetch_single(yahoo_ticker: str) -> dict:
    """Fetch return data for a single ticker."""
    res = fetch_returns_yahoo([yahoo_ticker])
    return res.get(yahoo_ticker, {p: None for p in PERIODS})


# ── Write back Excel ───────────────────────────────────────────────────────────
def write_excel(original_bytes: bytes, df: pd.DataFrame) -> bytes:
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(original_bytes))
    ws = wb["Current DR80"]

    raw = pd.read_excel(io.BytesIO(original_bytes), sheet_name="Current DR80", header=None)
    df_idx = df.set_index("BBG_Ticker")
    sector_last = {}
    current_sector = None
    original_bbg = set()

    for i, row in raw.iterrows():
        c1 = str(row[1]) if pd.notna(row[1]) else ""
        c2 = str(row[2]) if pd.notna(row[2]) else ""
        if c2 == "name" and c1 not in ["", "nan"] and pd.isna(row[0]):
            current_sector = c1.strip()
            continue
        if c1 in ["", "nan"] or c1.startswith("Unnamed"):
            continue
        bbg = c1.strip()
        original_bbg.add(bbg)
        if current_sector:
            sector_last[current_sector] = i + 1  # 1-indexed
        # Update return cells
        if bbg in df_idx.index:
            rec = df_idx.loc[bbg]
            for j, p in enumerate(PERIODS):
                v = rec[p]
                ws.cell(row=i + 1, column=5 + j).value = round(float(v), 6) if pd.notna(v) and v is not None else None

    # Append new pipeline rows
    new_rows = df[~df["BBG_Ticker"].isin(original_bbg) & ~df["Is_DR80"]]
    for _, rec in new_rows.iterrows():
        sector = rec["Sector"]
        insert_after = sector_last.get(sector, ws.max_row)
        ws.insert_rows(insert_after + 1)
        nr = insert_after + 1
        code = rec["BBG_Ticker"].rsplit(" ", 2)[0].strip()
        ws.cell(row=nr, column=1).value = code
        ws.cell(row=nr, column=2).value = rec["BBG_Ticker"]
        ws.cell(row=nr, column=3).value = rec["Name"]
        ws.cell(row=nr, column=4).value = rec["Quarter"]
        for j, p in enumerate(PERIODS):
            v = rec[p]
            ws.cell(row=nr, column=5 + j).value = round(float(v), 6) if pd.notna(v) and v is not None else None
        sector_last[sector] = nr  # stack subsequent inserts

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def graduate_to_dr80(df: pd.DataFrame, bbg_tickers: list) -> pd.DataFrame:
    """
    Promote pipeline securities to DR80 status in the DataFrame.
    Transforms e.g. 'MU US Equity' -> 'MU80 TB Equity', clears Quarter, sets Is_DR80=True.
    """
    df = df.copy()
    for bbg in bbg_tickers:
        mask = df["BBG_Ticker"] == bbg
        if not mask.any():
            continue
        # Build new DR80 ticker: take base code, append '80 TB Equity'
        code = bbg.rsplit(" ", 2)[0].strip()
        new_bbg = f"{code}80 TB Equity"
        df.loc[mask, "BBG_Ticker"] = new_bbg
        df.loc[mask, "Yahoo_Ticker"] = None   # TB tickers not on Yahoo
        df.loc[mask, "Is_DR80"] = True
        df.loc[mask, "Quarter"] = None
    return df


def write_excel_graduated(original_bytes: bytes, df: pd.DataFrame,
                           graduated: list) -> bytes:
    """
    Write Excel with graduation applied:
    - For each graduated ticker, find its pipeline row and update:
      col0 → NaN, col1 → new DR80 ticker, col3 → empty
    - Then call normal write_excel logic for return updates.
    """
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(original_bytes))
    ws = wb["Current DR80"]
    raw = pd.read_excel(io.BytesIO(original_bytes), sheet_name="Current DR80", header=None)

    # Build map: old_bbg -> new_bbg for graduated tickers
    grad_map = {}
    for old_bbg in graduated:
        code = old_bbg.rsplit(" ", 2)[0].strip()
        grad_map[old_bbg] = f"{code}80 TB Equity"

    for i, row in raw.iterrows():
        c1 = str(row[1]) if pd.notna(row[1]) else ""
        if c1.strip() in grad_map:
            new_bbg = grad_map[c1.strip()]
            xl_row = i + 1
            ws.cell(row=xl_row, column=1).value = None       # clear short code
            ws.cell(row=xl_row, column=2).value = new_bbg    # new DR80 ticker
            ws.cell(row=xl_row, column=4).value = None       # clear quarter

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    # Now run normal write_excel on top of the updated bytes
    return write_excel(buf.read(), df)


# ── Chart helpers ──────────────────────────────────────────────────────────────
C = {"bg": "rgba(0,0,0,0)", "grid": "#1e2d4a", "text": "#94a3b8",
     "pos": "#10b981", "neg": "#ef4444", "blue": "#3b82f6", "font": "IBM Plex Mono"}

def base_layout(h=350, margin=None):
    m = margin or dict(l=10, r=10, t=50, b=10)
    return dict(paper_bgcolor=C["bg"], plot_bgcolor=C["bg"],
                font=dict(family=C["font"], color=C["text"], size=14),
                height=h, margin=m)

def fmt_pct(v, d=1):
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "—"
    return f"+{v:.{d}f}%" if v >= 0 else f"{v:.{d}f}%"

def style_pct(v):
    if pd.isna(v): return "color:#475569"
    return f"color:{C['pos']}" if v >= 0 else f"color:{C['neg']}"

def bar_colors(vals):
    return [C["pos"] if v >= 0 else C["neg"] for v in vals]

@st.cache_data(show_spinner=False)
def build_top_bottom_chart(data: bytes, period: str):
    df = pd.read_json(io.BytesIO(data))
    df["S"] = df.apply(lambda r: display_label(r["BBG_Ticker"], r["Name"]), axis=1).astype(str)
    seen = {}; deduped = []
    for lbl in df["S"]:
        if lbl in seen: seen[lbl]+=1; deduped.append(f"{lbl} ({seen[lbl]})")
        else: seen[lbl]=0; deduped.append(lbl)
    df["S"] = deduped
    df = df.sort_values(period, ascending=False)
    half = min(15, len(df)//2)
    bar_df = pd.concat([df.head(half), df.tail(half)]).drop_duplicates().sort_values(period)
    bar_h = max(320, len(bar_df)*28+60)
    fig = go.Figure(go.Bar(
        x=bar_df[period], y=bar_df["S"].astype(str), orientation="h",
        marker_color=bar_colors(bar_df[period]), marker_line_width=0,
        text=[f"{v:+.1f}%" for v in bar_df[period]], textposition="outside",
        textfont=dict(family=C["font"], size=13, color=C["text"]),
        hovertemplate="<b>%{y}</b><br>%{x:.1f}%<extra></extra>",
    ))
    fig.add_vline(x=0, line_color="#334155", line_width=1)
    fig.update_layout(title=dict(text=f"Top & Bottom Performers — {period}", font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                      **base_layout(bar_h), xaxis=dict(showgrid=True,gridcolor=C["grid"],ticksuffix="%",tickfont=dict(size=14)),
                      yaxis=dict(showgrid=False,tickfont=dict(size=14),type="category"))
    return fig

@st.cache_data(show_spinner=False)
def build_heatmap(data: bytes, period: str):
    df = pd.read_json(io.BytesIO(data))
    df["S"] = df.apply(lambda r: display_label(r["BBG_Ticker"], r["Name"]), axis=1).astype(str)
    df["abs_ytd"] = df["YTD"].abs()
    df = df.dropna(subset=["YTD"]).nlargest(20, "abs_ytd")
    z = df[PERIODS].values
    heat_h = max(380, len(df)*26+60)
    text_vals = [[f"{v:+.0f}%" if not (v is None or np.isnan(v)) else "—" for v in row] for row in z]
    fig = go.Figure(go.Heatmap(
        z=z, x=PERIODS, y=df["S"].tolist(),
        colorscale=[[0.0,"#991b1b"],[0.3,"#7f1d1d"],[0.5,"#0f172a"],[0.7,"#064e3b"],[1.0,"#059669"]],
        zmid=0, text=text_vals, texttemplate="%{text}",
        textfont=dict(family=C["font"],size=13,color="#e2e8f0"),
        hovertemplate="<b>%{y}</b> — %{x}<br>%{z:.1f}%<extra></extra>",
        colorbar=dict(title=dict(text="Return %",font=dict(family=C["font"],size=13,color="#64748b")),
                      tickfont=dict(family=C["font"],size=13,color="#94a3b8"),
                      ticksuffix="%",thickness=14,len=0.9,tickvals=[-100,-50,0,50,100]),
    ))
    fig.update_layout(title=dict(text="Multi-Period Heatmap — Top 20 by |YTD|",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                      **base_layout(heat_h,margin=dict(l=10,r=80,t=50,b=10)),
                      xaxis=dict(showgrid=False,tickfont=dict(size=14,color="#94a3b8"),side="bottom"),
                      yaxis=dict(showgrid=False,autorange="reversed",tickfont=dict(size=14,color="#e2e8f0")))
    return fig


# ── Session state ──────────────────────────────────────────────────────────────
for key, default in [("df", None), ("excel_bytes", None),
                     ("last_refresh", None), ("source_label", None),
                     ("competitors_df", None), ("graduated", [])]:
    if key not in st.session_state:
        st.session_state[key] = default

# Auto-load default file
if st.session_state.df is None and os.path.exists(DEFAULT_FILE):
    with open(DEFAULT_FILE, "rb") as f:
        b = f.read()
    st.session_state.excel_bytes = b
    st.session_state.df = parse_excel(b)
    st.session_state.competitors_df = parse_competitors(b)
    st.session_state.source_label = DEFAULT_FILE


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.6rem;color:#334155;letter-spacing:0.1em;text-transform:uppercase;margin-bottom:16px;">DR80 DASHBOARD · KTB Securities</div>', unsafe_allow_html=True)

    # Upload
    st.markdown("**📂 DATA SOURCE**")
    uploaded = st.file_uploader("Upload Excel", type=["xlsx"], label_visibility="collapsed")
    if uploaded:
        b = uploaded.read()
        st.session_state.excel_bytes = b
        st.session_state.df = parse_excel(b)
        st.session_state.competitors_df = parse_competitors(b)
        st.session_state.source_label = uploaded.name
        st.session_state.last_refresh = None
        st.session_state.graduated = []
        comp_count = len(st.session_state.competitors_df)
        msg = f"✓ Loaded {uploaded.name}"
        if comp_count:
            msg += f" · {comp_count} competitors"
        st.success(msg)

    if st.session_state.source_label:
        st.caption(f"Source: {st.session_state.source_label}")
    if st.session_state.last_refresh:
        st.caption(f"Last refresh: {st.session_state.last_refresh}")

    st.markdown("---")

    # Refresh
    st.markdown("**🔄 REFRESH DATA**")
    st.caption("Fetches live returns from Yahoo Finance for all non-TB securities.")
    if st.button("⟳ Refresh from Yahoo Finance", use_container_width=True):
        if st.session_state.df is None:
            st.error("Load a file first.")
        else:
            df_curr = st.session_state.df.copy()
            fetchable = df_curr[df_curr["Yahoo_Ticker"].notna()][["BBG_Ticker", "Yahoo_Ticker"]].drop_duplicates()
            fetched = fetch_returns_yahoo(fetchable["Yahoo_Ticker"].tolist())
            ymap = dict(zip(fetchable["Yahoo_Ticker"], fetchable["BBG_Ticker"]))
            n = 0
            for yticker, rets in fetched.items():
                bbg = ymap.get(yticker)
                if bbg is None: continue
                mask = df_curr["BBG_Ticker"] == bbg
                for p in PERIODS:
                    if rets.get(p) is not None:
                        df_curr.loc[mask, p] = rets[p]
                n += 1
            st.session_state.df = df_curr
            st.session_state.last_refresh = datetime.now().strftime("%Y-%m-%d %H:%M")
            st.success(f"✓ Updated {n} securities")

    st.markdown("---")

    # Filters (only if data loaded)
    if st.session_state.df is not None:
        df_all = st.session_state.df

        st.markdown("**FILTERS**")
        universe = st.radio("Universe", ["All", "DR80 Only", "Pipeline Only"], label_visibility="collapsed")

        all_sectors = sorted(df_all["Sector"].unique())
        sel_sectors = st.multiselect("Sectors", all_sectors, default=all_sectors, label_visibility="collapsed",
                                     placeholder="All sectors")

        all_qtrs = sorted([q for q in df_all["Quarter"].dropna().unique()])
        sel_qtrs = st.multiselect("Pipeline quarters", all_qtrs, default=all_qtrs, label_visibility="collapsed",
                                  placeholder="All quarters")

        period = st.select_slider("Period", PERIODS, value="YTD", label_visibility="collapsed")

        valid_p = df_all[period].dropna()
        if len(valid_p):
            mn, mx = float(valid_p.min()), float(valid_p.max())
            ret_range = st.slider("Return range (%)", min_value=round(mn), max_value=round(mx),
                                  value=(round(mn), round(mx)), label_visibility="collapsed")
        else:
            ret_range = (-500, 1000)

        search = st.text_input("🔍 Search", placeholder="Ticker or name…", label_visibility="collapsed")


# ── No data ────────────────────────────────────────────────────────────────────
if st.session_state.df is None:
    st.markdown('<div class="dashboard-title">DR80 TRACKING DASHBOARD</div>', unsafe_allow_html=True)
    st.info("Upload a `DR80_Tracking.xlsx` file in the sidebar to get started.")
    st.stop()


# ── Apply filters ──────────────────────────────────────────────────────────────
df_all = st.session_state.df
filt = df_all.copy()

if universe == "DR80 Only":
    filt = filt[filt["Is_DR80"]]
elif universe == "Pipeline Only":
    filt = filt[~filt["Is_DR80"]]

if sel_sectors:
    filt = filt[filt["Sector"].isin(sel_sectors)]

# Quarter filter only applies to pipeline rows
filt = filt[filt["Is_DR80"] | filt["Quarter"].isin(sel_qtrs)]

filt = filt[filt[period].isna() | ((filt[period] >= ret_range[0]) & (filt[period] <= ret_range[1]))]

if search:
    s = search.lower()
    filt = filt[filt["BBG_Ticker"].str.lower().str.contains(s, na=False) |
                filt["Name"].str.lower().str.contains(s, na=False)]


# ══════════════════════════════════════════════════════════════════════════════
# TABS
# ══════════════════════════════════════════════════════════════════════════════
tab_dash, tab_sector, tab_pipeline, tab_mktshare, tab_add = st.tabs([
    "📊  Dashboard", "🔬  Sector Analysis", "🔭  Pipeline",
    "⚔️  Competitors & Market Share", "➕  Add Security",
])

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 1 — DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
with tab_dash:
    ct, ci = st.columns([3, 1])
    with ct:
        st.markdown('<div class="dashboard-title">DR80 TRACKING</div>', unsafe_allow_html=True)
        st.markdown('<div class="dashboard-sub">KTB SECURITIES · DEPOSITARY RECEIPT OPERATIONS</div>', unsafe_allow_html=True)
    with ci:
        st.markdown(f'<div style="text-align:right;font-family:IBM Plex Mono;font-size:0.65rem;color:#334155;margin-top:8px;">SHOWING {len(filt)} / {len(df_all)}<br>PERIOD: <span style="color:#3b82f6">{period}</span></div>', unsafe_allow_html=True)
    st.markdown("---")

    pv = filt[period].dropna()
    pos_n = int((pv >= 0).sum()); neg_n = int((pv < 0).sum())
    avg_r = float(pv.mean()) if len(pv) else 0.0
    hit_r = pos_n / len(pv) * 100 if len(pv) else 0.0
    best  = filt.loc[filt[period].idxmax()] if len(pv) else None
    worst = filt.loc[filt[period].idxmin()] if len(pv) else None

    k1, k2, k3, k4, k5 = st.columns(5)
    with k1:
        st.markdown(f"""<div class="metric-card"><div class="metric-label">Securities</div>
        <div class="metric-value">{len(filt)}</div>
        <div class="metric-sub">DR80: {filt['Is_DR80'].sum()} · Pipeline: {(~filt['Is_DR80']).sum()}</div></div>""", unsafe_allow_html=True)
    with k2:
        vc = C["pos"] if avg_r >= 0 else C["neg"]
        st.markdown(f"""<div class="metric-card {'green' if avg_r>=0 else 'red'}"><div class="metric-label">Avg Return ({period})</div>
        <div class="metric-value" style="color:{vc}">{fmt_pct(avg_r)}</div>
        <div class="metric-sub">{pos_n} pos · {neg_n} neg</div></div>""", unsafe_allow_html=True)
    with k3:
        bt = display_label(best["BBG_Ticker"], best["Name"]) if best is not None else "—"
        st.markdown(f"""<div class="metric-card green"><div class="metric-label">Best ({period})</div>
        <div class="metric-value" style="color:#10b981">{fmt_pct(best[period] if best is not None else None)}</div>
        <div class="metric-sub">{bt}</div></div>""", unsafe_allow_html=True)
    with k4:
        wt = display_label(worst["BBG_Ticker"], worst["Name"]) if worst is not None else "—"
        st.markdown(f"""<div class="metric-card red"><div class="metric-label">Worst ({period})</div>
        <div class="metric-value" style="color:#ef4444">{fmt_pct(worst[period] if worst is not None else None)}</div>
        <div class="metric-sub">{wt}</div></div>""", unsafe_allow_html=True)
    with k5:
        st.markdown(f"""<div class="metric-card"><div class="metric-label">Win Rate</div>
        <div class="metric-value">{hit_r:.0f}%</div>
        <div class="metric-sub">+ve return for {period}</div></div>""", unsafe_allow_html=True)

    st.markdown('<div class="section-header">BREAKDOWN</div>', unsafe_allow_html=True)
    pie1, pie2, pie3 = st.columns(3)

    with pie1:
        # By Sector donut
        sec_counts = filt["Sector"].value_counts()
        sec_colors_pie = ["#3b82f6","#10b981","#f59e0b","#8b5cf6","#ef4444","#06b6d4","#f97316","#84cc16"]
        fig_ps = go.Figure(go.Pie(
            labels=sec_counts.index, values=sec_counts.values, hole=0.55,
            marker=dict(colors=sec_colors_pie[:len(sec_counts)]),
            textfont=dict(family=C["font"], size=14),
            hovertemplate="<b>%{label}</b><br>%{value} securities<br>%{percent}<extra></extra>",
        ))
        fig_ps.update_layout(title=dict(text="By Sector", font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                             **base_layout(300), legend=dict(font=dict(family=C["font"],size=13,color="#94a3b8")))
        st.plotly_chart(fig_ps, use_container_width=True)

    with pie2:
        # By Type donut (DR80 vs Pipeline)
        type_counts = filt["Is_DR80"].value_counts()
        fig_pt = go.Figure(go.Pie(
            labels=["DR80" if k else "Pipeline" for k in type_counts.index],
            values=type_counts.values, hole=0.55,
            marker=dict(colors=["#3b82f6","#f59e0b"]),
            textfont=dict(family=C["font"], size=14),
            hovertemplate="<b>%{label}</b><br>%{value}<br>%{percent}<extra></extra>",
        ))
        fig_pt.update_layout(title=dict(text="DR80 vs Pipeline", font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                             **base_layout(300),
                             legend=dict(font=dict(family=C["font"],size=13,color="#94a3b8")),
                             annotations=[dict(text=f"<b>{len(filt)}</b><br>total", x=0.5, y=0.5,
                                               showarrow=False, font=dict(family=C["font"],size=16,color="#e2e8f0"))])
        st.plotly_chart(fig_pt, use_container_width=True)

    with pie3:
        # By Vintage (Quarter) donut
        vtg = filt["Quarter"].value_counts().sort_index()
        if len(vtg):
            fig_pv = go.Figure(go.Pie(
                labels=vtg.index, values=vtg.values, hole=0.55,
                marker=dict(colors=["#3b82f6","#10b981","#f59e0b","#8b5cf6"]),
                textfont=dict(family=C["font"], size=14),
                hovertemplate="<b>%{label}</b><br>%{value} securities<br>%{percent}<extra></extra>",
            ))
            fig_pv.update_layout(title=dict(text="By Vintage (Quarter)", font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                                 **base_layout(300), legend=dict(font=dict(family=C["font"],size=13,color="#94a3b8")))
            st.plotly_chart(fig_pv, use_container_width=True)
        else:
            st.info("No quarter data available.")

    # ── Vintage Analysis ───────────────────────────────────────────────────────
    st.markdown('<div class="section-header">VINTAGE ANALYSIS — PERFORMANCE BY ISSUANCE QUARTER</div>', unsafe_allow_html=True)
    vtg_df = filt.dropna(subset=["Quarter"]).copy()
    if len(vtg_df):
        vtg_df["Label"] = vtg_df.apply(lambda r: display_label(r["BBG_Ticker"],r["Name"]),axis=1).astype(str)

    if len(vtg_df) == 0:
        st.info("No quarter/vintage data available.")
    else:
        va1, va2 = st.columns([2, 3])
        with va1:
            # Bar: avg return per period per vintage
            vtg_avgs = vtg_df.groupby("Quarter")[PERIODS].mean()
            vtg_avgs = vtg_avgs.sort_index()
            fig_va = go.Figure()
            bar_colors_vtg = ["#3b82f6","#10b981","#f59e0b","#8b5cf6"]
            for i, (qtr, row) in enumerate(vtg_avgs.iterrows()):
                fig_va.add_trace(go.Bar(
                    name=qtr, x=PERIODS, y=row.values,
                    marker_color=bar_colors_vtg[i % len(bar_colors_vtg)],
                    text=[f"{v:+.0f}%" if not np.isnan(v) else "" for v in row.values],
                    textposition="outside",
                    textfont=dict(family=C["font"], size=13),
                    hovertemplate=f"<b>{qtr}</b><br>%{{x}}: %{{y:.1f}}%<extra></extra>",
                ))
            fig_va.add_hline(y=0, line_color="#334155", line_width=1)
            fig_va.update_layout(
                title=dict(text="Avg Return by Vintage & Period", font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                **base_layout(360, margin=dict(l=10,r=10,t=50,b=10)),
                barmode="group",
                xaxis=dict(showgrid=False, tickfont=dict(size=14)),
                yaxis=dict(showgrid=True, gridcolor=C["grid"], ticksuffix="%", tickfont=dict(size=14)),
                legend=dict(font=dict(family=C["font"],size=13,color="#94a3b8"), orientation="h", y=1.12),
            )
            st.plotly_chart(fig_va, use_container_width=True)

        with va2:
            # Heatmap: each security row, sorted by quarter then YTD
            vtg_heat = vtg_df.sort_values(["Quarter","YTD"], ascending=[True, False])
            vtg_heat = vtg_heat[["Label","Quarter"]+PERIODS].dropna(subset=["YTD"]).head(20)
            # Label includes quarter prefix for readability
            vtg_heat["RowLabel"] = vtg_heat["Quarter"].astype(str) + "  " + vtg_heat["Label"].astype(str)
            z_v = vtg_heat[PERIODS].values.astype(float)
            text_v = [[f"{v:+.0f}%" if not np.isnan(v) else "—" for v in row] for row in z_v]
            fig_vh = go.Figure(go.Heatmap(
                z=z_v, x=PERIODS, y=vtg_heat["RowLabel"].tolist(),
                colorscale=[[0.0,"#991b1b"],[0.3,"#7f1d1d"],[0.5,"#0f172a"],[0.7,"#064e3b"],[1.0,"#059669"]],
                zmid=0, text=text_v, texttemplate="%{text}",
                textfont=dict(family=C["font"], size=16, color="#e2e8f0"),
                hovertemplate="<b>%{y}</b> — %{x}<br>%{z:.1f}%<extra></extra>",
                colorbar=dict(tickfont=dict(family=C["font"],size=13,color="#94a3b8"), ticksuffix="%",
                              thickness=14, len=0.9,
                              title=dict(text="Return %",font=dict(family=C["font"],size=13,color="#64748b"))),
            ))
            fig_vh.update_layout(
                title=dict(text="Vintage Heatmap — Top 30 by Quarter", font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                **base_layout(max(400, len(vtg_heat)*26+60), margin=dict(l=10,r=80,t=50,b=10)),
                xaxis=dict(showgrid=False, tickfont=dict(size=14,color="#94a3b8")),
                yaxis=dict(showgrid=False, autorange="reversed", tickfont=dict(size=13,color="#e2e8f0")),
            )
            st.plotly_chart(fig_vh, use_container_width=True)

    st.markdown('<div class="section-header">PERFORMANCE</div>', unsafe_allow_html=True)
    ca, cb = st.columns([3, 2])

    with ca:
        pdf = filt[["BBG_Ticker","Name",period]].dropna(subset=[period]).copy()
        pdf["S"] = pdf.apply(lambda r: display_label(r["BBG_Ticker"], r["Name"]), axis=1).astype(str)
        seen = {}; deduped = []
        for lbl in pdf["S"]:
            if lbl in seen: seen[lbl]+=1; deduped.append(f"{lbl} ({seen[lbl]})")
            else: seen[lbl]=0; deduped.append(lbl)
        pdf["S"] = deduped
        pdf = pdf.sort_values(period, ascending=False)
        half = min(15, len(pdf)//2)
        bar_df = pd.concat([pdf.head(half), pdf.tail(half)]).drop_duplicates().sort_values(period)
        bar_h = max(320, len(bar_df)*28+60)
        fig = go.Figure(go.Bar(
            x=bar_df[period], y=bar_df["S"].astype(str), orientation="h",
            marker_color=bar_colors(bar_df[period]), marker_line_width=0,
            text=[f"{v:+.1f}%" for v in bar_df[period]], textposition="outside",
            textfont=dict(family=C["font"], size=13, color=C["text"]),
            hovertemplate="<b>%{y}</b><br>%{x:.1f}%<extra></extra>",
        ))
        fig.add_vline(x=0, line_color="#334155", line_width=1)
        fig.update_layout(title=dict(text=f"Top & Bottom Performers — {period}", font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                          **base_layout(bar_h), xaxis=dict(showgrid=True,gridcolor=C["grid"],ticksuffix="%",tickfont=dict(size=14)),
                          yaxis=dict(showgrid=False,tickfont=dict(size=14),type="category"))
        st.plotly_chart(fig, use_container_width=True)

    with cb:
        sp = filt.groupby("Sector")[period].mean().dropna().sort_values()
        fig2 = go.Figure(go.Bar(
            x=sp.values, y=sp.index, orientation="h",
            marker_color=bar_colors(sp.values), marker_line_width=0,
            text=[f"{v:+.1f}%" for v in sp.values], textposition="outside",
            textfont=dict(family=C["font"],size=13,color=C["text"]),
            hovertemplate="<b>%{y}</b><br>%{x:.1f}%<extra></extra>",
        ))
        fig2.add_vline(x=0, line_color="#334155", line_width=1)
        fig2.update_layout(title=dict(text=f"Avg by Sector — {period}",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                           **base_layout(420,margin=dict(l=10,r=80,t=44,b=10)),
                           xaxis=dict(showgrid=True,gridcolor=C["grid"],ticksuffix="%",tickfont=dict(size=14)),
                           yaxis=dict(showgrid=False,tickfont=dict(size=14)))
        st.plotly_chart(fig2, use_container_width=True)

    st.markdown('<div class="section-header">DISTRIBUTION & HEATMAPS</div>', unsafe_allow_html=True)
    cc1, cc2 = st.columns([2, 3])

    with cc1:
        hv = filt[period].dropna()
        fig3 = go.Figure(go.Histogram(x=hv, nbinsx=20, marker_color=C["blue"],
                                      marker_line_color=C["grid"], marker_line_width=1, opacity=0.85))
        fig3.add_vline(x=0, line_color=C["neg"], line_width=1, line_dash="dash")
        if len(hv):
            fig3.add_vline(x=float(hv.mean()), line_color="#f59e0b", line_width=1, line_dash="dot",
                           annotation_text=f"avg {hv.mean():+.1f}%",
                           annotation_font=dict(color="#f59e0b",size=10,family=C["font"]))
        fig3.update_layout(title=dict(text=f"Distribution — {period}",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                           **base_layout(340),xaxis=dict(showgrid=True,gridcolor=C["grid"],ticksuffix="%",tickfont=dict(size=14)),
                           yaxis=dict(showgrid=True,gridcolor=C["grid"],tickfont=dict(size=14)))
        st.plotly_chart(fig3, use_container_width=True)

    with cc2:
        hdf = filt[["BBG_Ticker","Name"]+PERIODS].copy()
        hdf["S"] = hdf.apply(lambda r: display_label(r["BBG_Ticker"],r["Name"]),axis=1).astype(str)
        hdf["abs_ytd"] = hdf["YTD"].abs()
        hdf = hdf.dropna(subset=["YTD"]).nlargest(20,"abs_ytd")
        z = hdf[PERIODS].values
        heat_h = max(380, len(hdf)*26+60)
        text_vals = [[f"{v:+.0f}%" if not (v is None or np.isnan(v)) else "—" for v in row] for row in z]
        fig4 = go.Figure(go.Heatmap(
            z=z, x=PERIODS, y=hdf["S"].tolist(),
            colorscale=[[0.0,"#991b1b"],[0.3,"#7f1d1d"],[0.5,"#0f172a"],[0.7,"#064e3b"],[1.0,"#059669"]],
            zmid=0, text=text_vals, texttemplate="%{text}",
            textfont=dict(family=C["font"],size=16,color="#e2e8f0"),
            hovertemplate="<b>%{y}</b> — %{x}<br>%{z:.1f}%<extra></extra>",
            colorbar=dict(title=dict(text="Return %",font=dict(family=C["font"],size=13,color="#64748b")),
                          tickfont=dict(family=C["font"],size=13,color="#94a3b8"),
                          ticksuffix="%",thickness=14,len=0.9,tickvals=[-100,-50,0,50,100]),
        ))
        fig4.update_layout(title=dict(text="Multi-Period Heatmap — Top 20 by |YTD|",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                           **base_layout(heat_h,margin=dict(l=10,r=80,t=44,b=10)),
                           xaxis=dict(showgrid=False,tickfont=dict(size=12,color="#94a3b8"),side="bottom"),
                           yaxis=dict(showgrid=False,autorange="reversed",tickfont=dict(size=11,color="#e2e8f0")))
        st.plotly_chart(fig4, use_container_width=True)

    st.markdown('<div class="section-header">SECTOR HEATMAP — AVG & MEDIAN RETURNS</div>', unsafe_allow_html=True)
    _sec_avgs = filt.groupby("Sector")[PERIODS].mean()
    _sec_meds = filt.groupby("Sector")[PERIODS].median()
    z_sec=[]; y_sec=[]
    for sec in sorted(filt["Sector"].unique()):
        z_sec.append(_sec_avgs.loc[sec].values if sec in _sec_avgs.index else [np.nan]*len(PERIODS)); y_sec.append(f"{sec}  avg")
        z_sec.append(_sec_meds.loc[sec].values if sec in _sec_meds.index else [np.nan]*len(PERIODS)); y_sec.append(f"{sec}  med")
    z_sec_arr = np.array(z_sec, dtype=float)
    sec_h = max(300, len(y_sec)*22+60)
    text_sec = [[f"{v:+.0f}%" if not np.isnan(v) else "—" for v in row] for row in z_sec_arr]
    fig_sh2 = go.Figure(go.Heatmap(
        z=z_sec_arr, x=PERIODS, y=y_sec,
        colorscale=[[0.0,"#991b1b"],[0.3,"#7f1d1d"],[0.5,"#0f172a"],[0.7,"#064e3b"],[1.0,"#059669"]],
        zmid=0, text=text_sec, texttemplate="%{text}",
        textfont=dict(family=C["font"],size=16,color="#e2e8f0"),
        hovertemplate="<b>%{y}</b> — %{x}<br>%{z:.1f}%<extra></extra>",
        colorbar=dict(title=dict(text="Return %",font=dict(family=C["font"],size=13,color="#64748b")),
                      tickfont=dict(family=C["font"],size=13,color="#94a3b8"),
                      ticksuffix="%",thickness=14,len=0.9,tickvals=[-50,-25,0,25,50]),
    ))
    fig_sh2.update_layout(title=dict(text="Sector Returns by Period (Avg & Median)",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                          **base_layout(sec_h,margin=dict(l=10,r=80,t=44,b=10)),
                          xaxis=dict(showgrid=False,tickfont=dict(size=12,color="#94a3b8")),
                          yaxis=dict(showgrid=False,autorange="reversed",tickfont=dict(size=11,color="#e2e8f0")))
    st.plotly_chart(fig_sh2, use_container_width=True)

    # ── PRICE RETURN BY DURATION ───────────────────────────────────────────────
    st.markdown('<div class="section-header">PRICE RETURN BY DURATION</div>', unsafe_allow_html=True)
    st.caption("Compare every security's return across all time periods at once.")

    prd_df = filt[["BBG_Ticker","Name","Sector","Is_DR80"] + PERIODS].copy()
    prd_df["Ticker"] = prd_df.apply(lambda r: display_label(r["BBG_Ticker"], r["Name"]), axis=1).astype(str)
    prd_df["Type"]   = prd_df["Is_DR80"].map({True: "DR80", False: "Pipeline"})
    prd_df = prd_df.drop(columns=["BBG_Ticker","Is_DR80"])

    prd_c1, prd_c2, prd_c3 = st.columns([2, 2, 2])
    with prd_c1:
        prd_type = st.radio("Show", ["All", "DR80 Only", "Pipeline Only"],
                            horizontal=True, key="prd_type")
    with prd_c2:
        prd_sort = st.selectbox("Sort by period", PERIODS, index=PERIODS.index("YTD"),
                                key="prd_sort", label_visibility="collapsed")
    with prd_c3:
        prd_asc = st.checkbox("Ascending", value=False, key="prd_asc")

    if prd_type == "DR80 Only":
        prd_df = prd_df[prd_df["Type"] == "DR80"]
    elif prd_type == "Pipeline Only":
        prd_df = prd_df[prd_df["Type"] == "Pipeline"]

    prd_df = prd_df.sort_values(prd_sort, ascending=prd_asc, na_position="last").reset_index(drop=True)

    # Per-period KPI summary row
    kpi_cols = st.columns(len(PERIODS))
    for _ki, _p in enumerate(PERIODS):
        _col_data = prd_df[_p].dropna()
        if len(_col_data) == 0:
            kpi_cols[_ki].markdown(f'<div class="metric-card"><div class="metric-label">{_p}</div><div class="metric-value">—</div></div>', unsafe_allow_html=True)
            continue
        _avg = _col_data.mean()
        _vc = C["pos"] if _avg >= 0 else C["neg"]
        _cls = "green" if _avg >= 0 else "red"
        _best_nm  = prd_df.loc[_col_data.idxmax(), "Ticker"] if _col_data.idxmax() in prd_df.index else "—"
        _worst_nm = prd_df.loc[_col_data.idxmin(), "Ticker"] if _col_data.idxmin() in prd_df.index else "—"
        kpi_cols[_ki].markdown(f"""
<div class="metric-card {_cls}" style="padding:10px 14px;">
  <div class="metric-label">{_p}</div>
  <div class="metric-value" style="color:{_vc};font-size:1.3rem;">{fmt_pct(_avg)}</div>
  <div class="metric-sub" style="color:#10b981;font-size:0.7rem;">▲ {fmt_pct(_col_data.max())} {_best_nm}</div>
  <div class="metric-sub" style="color:#ef4444;font-size:0.7rem;">▼ {fmt_pct(_col_data.min())} {_worst_nm}</div>
</div>""", unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    # Grouped bar: top+bottom N, all periods side by side
    prd_top_n = st.slider("Top / bottom N", 5, 25, 12, key="prd_topn")
    _top_h = prd_df.dropna(subset=[prd_sort]).head(prd_top_n)
    _bot_h = prd_df.dropna(subset=[prd_sort]).tail(prd_top_n)
    prd_show = pd.concat([_top_h, _bot_h]).drop_duplicates().reset_index(drop=True)

    period_colors_map = {
        "YTD":"#3b82f6","1M":"#10b981","3M":"#f59e0b",
        "6M":"#8b5cf6","1Y":"#06b6d4","3Y":"#f97316","5Y":"#ef4444"
    }
    fig_dur = go.Figure()
    for _p in PERIODS:
        fig_dur.add_trace(go.Bar(
            name=_p,
            x=prd_show["Ticker"].astype(str),
            y=prd_show[_p].fillna(0),
            marker_color=period_colors_map.get(_p,"#64748b"),
            opacity=0.85,
            hovertemplate=f"<b>%{{x}}</b> — {_p}<br>%{{y:.1f}}%<extra></extra>",
        ))
    fig_dur.add_hline(y=0, line_color="#334155", line_width=1)
    fig_dur.update_layout(
        title=dict(text=f"Multi-Period Returns — Top & Bottom {prd_top_n} by {prd_sort}",
                   font=dict(family=C["font"], size=15, color="#64748b"), x=0),
        **base_layout(420, margin=dict(l=10, r=10, t=50, b=80)),
        barmode="group",
        xaxis=dict(showgrid=False, tickfont=dict(size=10), tickangle=-35),
        yaxis=dict(showgrid=True, gridcolor=C["grid"], ticksuffix="%", tickfont=dict(size=12)),
        legend=dict(font=dict(family=C["font"], size=12, color="#94a3b8"),
                    orientation="h", y=1.08, x=0),
    )
    st.plotly_chart(fig_dur, use_container_width=True)

    # Momentum score bar
    st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.7rem;color:#475569;text-transform:uppercase;letter-spacing:0.1em;margin:8px 0 10px;">MOMENTUM SCORE — # PERIODS IN POSITIVE TERRITORY</div>', unsafe_allow_html=True)
    prd_df["Momentum"] = prd_df[PERIODS].apply(lambda row: int((row > 0).sum()), axis=1)
    mom_sorted = prd_df.sort_values("Momentum", ascending=False).head(20)
    fig_mom = go.Figure(go.Bar(
        x=mom_sorted["Ticker"].astype(str),
        y=mom_sorted["Momentum"],
        marker_color=["#3b82f6" if t=="DR80" else "#f59e0b" for t in mom_sorted["Type"]],
        marker_line_width=0,
        text=[f"{v}/{len(PERIODS)}" for v in mom_sorted["Momentum"]],
        textposition="outside",
        textfont=dict(family=C["font"], size=11, color=C["text"]),
        hovertemplate="<b>%{x}</b><br>%{y} of " + str(len(PERIODS)) + " periods positive<extra></extra>",
    ))
    fig_mom.update_layout(
        title=dict(text="Momentum Score — Periods in Positive Territory (Top 20)",
                   font=dict(family=C["font"], size=14, color="#64748b"), x=0),
        **base_layout(300, margin=dict(l=10, r=10, t=44, b=60)),
        xaxis=dict(showgrid=False, tickfont=dict(size=10), tickangle=-35),
        yaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=12),
                   range=[0, len(PERIODS) + 0.5]),
        showlegend=False,
    )
    st.plotly_chart(fig_mom, use_container_width=True)

    # Full expandable duration table
    with st.expander("📋 Full Duration Table", expanded=False):
        dur_tbl = prd_df[["Ticker","Name","Sector","Type","Momentum"] + PERIODS].copy()
        st.dataframe(
            dur_tbl.style
                .applymap(style_pct, subset=PERIODS)
                .format({p: lambda x: fmt_pct(x) for p in PERIODS})
                .set_properties(**{"font-family":"IBM Plex Mono","font-size":"12px"}),
            use_container_width=True, height=450
        )

    st.markdown('<div class="section-header">SECURITY TABLE</div>', unsafe_allow_html=True)
    sc1, sc2 = st.columns([2,1])
    with sc1: sort_by = st.selectbox("Sort by", ["Name","Sector"]+PERIODS, index=2, label_visibility="collapsed")
    with sc2: asc = st.checkbox("Ascending", value=False)
    tbl = filt[["BBG_Ticker","Name","Sector","Is_DR80","Quarter"]+PERIODS].copy()
    tbl["Ticker"] = tbl.apply(lambda r: display_label(r["BBG_Ticker"],r["Name"]),axis=1).astype(str)
    tbl["Type"]   = tbl["Is_DR80"].map({True:"DR80",False:"Pipeline"})
    tbl = tbl.drop(columns=["BBG_Ticker","Is_DR80"]).rename(columns={"Quarter":"Q"})
    tbl = tbl[["Ticker","Name","Sector","Type","Q"]+PERIODS].sort_values(sort_by,ascending=asc,na_position="last")
    st.dataframe(tbl.style.applymap(style_pct,subset=PERIODS)
                 .format({p: lambda x: fmt_pct(x) for p in PERIODS})
                 .set_properties(**{"font-family":"IBM Plex Mono","font-size":"12px"}),
                 use_container_width=True, height=430)
    e1, e2 = st.columns(2)
    with e1:
        st.download_button("⬇ Export CSV", data=tbl.to_csv(index=False),
                           file_name=f"DR80_{period}_{datetime.now().strftime('%Y%m%d')}.csv",
                           mime="text/csv", use_container_width=True)
    with e2:
        if st.session_state.excel_bytes:
            xl_out = write_excel(st.session_state.excel_bytes, st.session_state.df)
            st.download_button("⬇ Export Updated Excel", data=xl_out,
                               file_name=f"DR80_Tracking_{datetime.now().strftime('%Y%m%d')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                               use_container_width=True)
# ═══════════════════════════════════════════════════════════════════════════════
with tab_sector:
    st.markdown('<div class="section-header">SECTOR DEEP-DIVE</div>', unsafe_allow_html=True)
    sel_sector = st.selectbox("Select sector", sorted(df_all["Sector"].unique()),
                              label_visibility="collapsed", key="sector_drill")
    sec_df = df_all[df_all["Sector"]==sel_sector].copy()

    # DR80 / Pipeline filter
    sec_universe = st.radio("Universe", ["All", "DR80 Only", "Pipeline Only"],
                            horizontal=True, label_visibility="collapsed", key="sector_universe")
    if sec_universe == "DR80 Only":
        sec_df = sec_df[sec_df["Is_DR80"]]
    elif sec_universe == "Pipeline Only":
        sec_df = sec_df[~sec_df["Is_DR80"]]

    sec_df["Label"] = sec_df.apply(lambda r: display_label(r["BBG_Ticker"],r["Name"]),axis=1).astype(str)

    sv = sec_df["YTD"].dropna()
    s1,s2,s3,s4 = st.columns(4)
    with s1:
        st.markdown(f"""<div class="metric-card"><div class="metric-label">Securities</div>
        <div class="metric-value">{len(sec_df)}</div>
        <div class="metric-sub">DR80: {sec_df['Is_DR80'].sum()} · Pipeline: {(~sec_df['Is_DR80']).sum()}</div></div>""", unsafe_allow_html=True)
    with s2:
        avg_s = float(sv.mean()) if len(sv) else 0
        vc2 = C["pos"] if avg_s>=0 else C["neg"]
        st.markdown(f"""<div class="metric-card {'green' if avg_s>=0 else 'red'}"><div class="metric-label">Avg YTD Return</div>
        <div class="metric-value" style="color:{vc2}">{fmt_pct(avg_s)}</div>
        <div class="metric-sub">Median: {fmt_pct(float(sv.median()) if len(sv) else None)}</div></div>""", unsafe_allow_html=True)
    with s3:
        if len(sv):
            bs = sec_df.loc[sec_df["YTD"].idxmax()]
            st.markdown(f"""<div class="metric-card green"><div class="metric-label">Best YTD</div>
            <div class="metric-value" style="color:#10b981">{fmt_pct(bs['YTD'])}</div>
            <div class="metric-sub">{bs['Label']}</div></div>""", unsafe_allow_html=True)
    with s4:
        if len(sv):
            ws2 = sec_df.loc[sec_df["YTD"].idxmin()]
            st.markdown(f"""<div class="metric-card red"><div class="metric-label">Worst YTD</div>
            <div class="metric-value" style="color:#ef4444">{fmt_pct(ws2['YTD'])}</div>
            <div class="metric-sub">{ws2['Label']}</div></div>""", unsafe_allow_html=True)

    st.markdown('<div class="section-header">RETURNS BY SECURITY</div>', unsafe_allow_html=True)
    sec_period = st.select_slider("Period", PERIODS, value="YTD", label_visibility="collapsed", key="sector_period")
    sc_a, sc_b = st.columns([3,2])

    with sc_a:
        bar_sec = sec_df[["Label","Name",sec_period]].dropna(subset=[sec_period]).sort_values(sec_period)
        fig_sb = go.Figure(go.Bar(
            x=bar_sec[sec_period], y=bar_sec["Label"].astype(str), orientation="h",
            marker_color=bar_colors(bar_sec[sec_period]), marker_line_width=0,
            text=[f"{v:+.1f}%" for v in bar_sec[sec_period]], textposition="outside",
            textfont=dict(family=C["font"],size=13,color=C["text"]),
            hovertemplate="<b>%{y}</b><br>%{x:.1f}%<extra></extra>",
        ))
        fig_sb.add_vline(x=0, line_color="#334155", line_width=1)
        fig_sb.update_layout(title=dict(text=f"{sel_sector} — {sec_period}",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                             **base_layout(max(300,len(bar_sec)*30+60)),
                             xaxis=dict(showgrid=True,gridcolor=C["grid"],ticksuffix="%",tickfont=dict(size=14)),
                             yaxis=dict(showgrid=False,tickfont=dict(size=14),type="category"))
        st.plotly_chart(fig_sb, use_container_width=True)

    with sc_b:
        pt1, pt2 = st.tabs(["DR80 vs Pipeline", "Quarter Split"])
        with pt1:
            type_counts = sec_df["Is_DR80"].value_counts()
            fp1 = go.Figure(go.Pie(
                labels=["DR80" if k else "Pipeline" for k in type_counts.index],
                values=type_counts.values, hole=0.55,
                marker_colors=["#3b82f6","#f59e0b"],
                textfont=dict(family=C["font"],size=14),
                hovertemplate="<b>%{label}</b><br>%{value}<br>%{percent}<extra></extra>",
            ))
            fp1.update_layout(**base_layout(280),
                              legend=dict(font=dict(family=C["font"],size=13,color="#64748b")),
                              annotations=[dict(text=f"<b>{len(sec_df)}</b><br>total",x=0.5,y=0.5,
                                                showarrow=False,font=dict(family=C["font"],size=16,color="#e2e8f0"))])
            st.plotly_chart(fp1, use_container_width=True)
        with pt2:
            qc2 = sec_df["Quarter"].value_counts().sort_index()
            if len(qc2):
                fp2 = go.Figure(go.Pie(labels=qc2.index, values=qc2.values, hole=0.55,
                                       marker_colors=["#3b82f6","#10b981","#f59e0b","#8b5cf6"],
                                       textfont=dict(family=C["font"],size=14),
                                       hovertemplate="<b>%{label}</b><br>%{value}<extra></extra>"))
                fp2.update_layout(**base_layout(280),legend=dict(font=dict(family=C["font"],size=13,color="#64748b")))
                st.plotly_chart(fp2, use_container_width=True)
            else:
                st.info("No pipeline securities in this sector.")

    st.markdown('<div class="section-header">MULTI-PERIOD HEATMAP</div>', unsafe_allow_html=True)
    heat_sec = sec_df[["Label"]+PERIODS].dropna(subset=["YTD"]).copy()
    z_s = heat_sec[PERIODS].values
    text_hs = [[f"{v:+.0f}%" if not(v is None or np.isnan(v)) else "—" for v in row] for row in z_s]
    fig_sh = go.Figure(go.Heatmap(
        z=z_s, x=PERIODS, y=heat_sec["Label"].astype(str).tolist(),
        colorscale=[[0.0,"#991b1b"],[0.3,"#7f1d1d"],[0.5,"#0f172a"],[0.7,"#064e3b"],[1.0,"#059669"]],
        zmid=0, text=text_hs, texttemplate="%{text}",
        textfont=dict(family=C["font"],size=16,color="#e2e8f0"),
        hovertemplate="<b>%{y}</b> — %{x}<br>%{z:.1f}%<extra></extra>",
        colorbar=dict(tickfont=dict(family=C["font"],size=13,color="#94a3b8"),ticksuffix="%",thickness=14,len=0.9,
                      title=dict(text="Return %",font=dict(family=C["font"],size=13,color="#64748b"))),
    ))
    fig_sh.update_layout(title=dict(text=f"{sel_sector} — All Periods",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                         **base_layout(max(260,len(heat_sec)*28+60),margin=dict(l=10,r=80,t=44,b=10)),
                         xaxis=dict(showgrid=False,tickfont=dict(size=14)),
                         yaxis=dict(showgrid=False,autorange="reversed",tickfont=dict(size=14)))
    st.plotly_chart(fig_sh, use_container_width=True)

    st.markdown('<div class="section-header">SECURITY TABLE</div>', unsafe_allow_html=True)
    stbl = sec_df[["Label","Name","Is_DR80","Quarter"]+PERIODS].copy()
    stbl["Type"] = stbl["Is_DR80"].map({True:"DR80",False:"Pipeline"})
    stbl = stbl.drop(columns=["Is_DR80"]).rename(columns={"Label":"Ticker","Quarter":"Q"})
    stbl = stbl[["Ticker","Name","Type","Q"]+PERIODS]
    st.dataframe(stbl.style.applymap(style_pct,subset=PERIODS)
                 .format({p: lambda x: fmt_pct(x) for p in PERIODS})
                 .set_properties(**{"font-family":"IBM Plex Mono","font-size":"12px"}),
                 use_container_width=True, height=400)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 3 — PIPELINE
# ═══════════════════════════════════════════════════════════════════════════════
with tab_pipeline:
    # ── Graduate to DR80 ───────────────────────────────────────────────────────
    st.markdown('<div class="section-header">🎓 GRADUATE TO DR80</div>', unsafe_allow_html=True)
    st.caption("Promote launched pipeline securities to DR80 status. Ticker converts automatically (e.g. MU US Equity → MU80 TB Equity). Q1 pre-selected as the upcoming launch cohort.")

    all_pipe = df_all[~df_all["Is_DR80"]].copy()
    if len(all_pipe) == 0:
        st.info("No pipeline securities available to graduate.")
    else:
        all_pipe["Display"] = all_pipe.apply(
            lambda r: f"[{r['Quarter'] or '?'}]  {display_label(r['BBG_Ticker'], r['Name'])}  —  {r['Name'][:40]}",
            axis=1
        )
        q1_tickers = all_pipe[all_pipe["Quarter"] == "Q1"]["BBG_Ticker"].tolist()
        all_options = all_pipe["BBG_Ticker"].tolist()
        all_displays = dict(zip(all_pipe["BBG_Ticker"], all_pipe["Display"]))

        grad_col1, grad_col2 = st.columns([3, 1])
        with grad_col1:
            to_graduate = st.multiselect(
                "Select securities to graduate",
                options=all_options,
                default=q1_tickers,
                format_func=lambda x: all_displays.get(x, x),
                label_visibility="collapsed",
                key="grad_select",
                placeholder="Select securities to promote to DR80..."
            )
        with grad_col2:
            if to_graduate:
                st.markdown(f"""<div class="metric-card green" style="margin-top:4px;">
                <div class="metric-label">Selected</div>
                <div class="metric-value" style="color:#10b981">{len(to_graduate)}</div>
                <div class="metric-sub">ready to graduate</div></div>""", unsafe_allow_html=True)

        if to_graduate:
            # Ticker preview
            preview_html = "".join([
                f'<span style="color:#64748b;font-size:0.8rem;">{b.rsplit(" ",2)[0]}  <span style="color:#3b82f6">→</span>  <span style="color:#10b981">{b.rsplit(" ",2)[0]}80 TB Equity</span></span><br>'
                for b in to_graduate[:6]
            ])
            if len(to_graduate) > 6:
                preview_html += f'<span style="color:#334155;font-size:0.75rem;">+ {len(to_graduate)-6} more</span>'
            st.markdown(f'<div style="font-family:IBM Plex Mono;background:#0d1221;border:1px solid #1e2d4a;border-radius:6px;padding:10px 14px;margin-bottom:12px;">{preview_html}</div>', unsafe_allow_html=True)

            btn1, btn2, _ = st.columns([1.2, 1.5, 2])
            with btn1:
                if st.button("🎓  Graduate to DR80", use_container_width=True, type="primary", key="grad_btn"):
                    st.session_state.df = graduate_to_dr80(st.session_state.df, to_graduate)
                    st.session_state.graduated = st.session_state.get("graduated", []) + to_graduate
                    st.success(f"✓ {len(to_graduate)} securities promoted to DR80. Download the updated Excel to save permanently.")
                    st.rerun()
            with btn2:
                if st.session_state.excel_bytes and st.session_state.get("graduated"):
                    try:
                        xl_grad = write_excel_graduated(
                            st.session_state.excel_bytes,
                            st.session_state.df,
                            st.session_state.graduated
                        )
                        st.download_button(
                            "⬇  Download with Graduations",
                            data=xl_grad,
                            file_name=f"DR80_Tracking_graduated_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                            key="grad_download"
                        )
                    except Exception as e:
                        st.error(f"Excel write failed: {e}")

    st.markdown("---")
    pipe_df_all = df_all[~df_all["Is_DR80"]].copy()
    st.markdown('<div class="section-header">PIPELINE OVERVIEW</div>', unsafe_allow_html=True)
    if len(pipe_df_all) == 0:
        st.info("No pipeline securities found.")
    else:
        pipe_sectors = sorted(pipe_df_all["Sector"].unique())
        pipe_sel_sectors = st.multiselect("Filter by sector", options=pipe_sectors, default=pipe_sectors,
                                          label_visibility="collapsed", placeholder="All sectors",
                                          key="pipe_sector_filter")
        pipe_df = pipe_df_all[pipe_df_all["Sector"].isin(pipe_sel_sectors)] if pipe_sel_sectors else pipe_df_all
        st.caption(f"Showing {len(pipe_df)} of {len(pipe_df_all)} pipeline securities")

        p1,p2,p3 = st.columns(3)
        with p1:
            qc = pipe_df["Quarter"].value_counts().sort_index()
            fp = go.Figure(go.Pie(labels=qc.index, values=qc.values, hole=0.6,
                                  marker_colors=["#3b82f6","#10b981","#f59e0b","#8b5cf6"],
                                  textfont=dict(family=C["font"],size=14),
                                  hovertemplate="<b>%{label}</b><br>%{value}<extra></extra>"))
            fp.update_layout(**base_layout(280),legend=dict(font=dict(family=C["font"],size=13,color="#64748b")),
                             title=dict(text="By Quarter",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                             annotations=[dict(text=f"<b>{len(pipe_df)}</b><br>total",x=0.5,y=0.5,
                                               showarrow=False,font=dict(family=C["font"],size=16,color="#e2e8f0"))])
            st.plotly_chart(fp, use_container_width=True)
        with p2:
            sc = pipe_df["Sector"].value_counts()
            fs = go.Figure(go.Pie(labels=sc.index, values=sc.values, hole=0.5,
                                  textfont=dict(family=C["font"],size=13),
                                  hovertemplate="<b>%{label}</b><br>%{value}<extra></extra>"))
            fs.update_layout(**base_layout(280),legend=dict(font=dict(family=C["font"],size=13,color="#64748b")),
                             title=dict(text="By Sector",font=dict(family=C["font"],size=15,color="#64748b"),x=0))
            st.plotly_chart(fs, use_container_width=True)
        with p3:
            qa = pipe_df.groupby("Quarter")["YTD"].mean().dropna().sort_index()
            fq = go.Figure(go.Bar(x=qa.index, y=qa.values, marker_color=bar_colors(qa.values),
                                  text=[f"{v:+.1f}%" for v in qa.values], textposition="outside",
                                  textfont=dict(family=C["font"],size=13,color=C["text"])))
            fq.add_hline(y=0, line_color="#334155", line_width=1)
            fq.update_layout(**base_layout(280,margin=dict(l=10,r=10,t=44,b=30)),
                             title=dict(text="Avg YTD by Quarter",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                             xaxis=dict(showgrid=False,tickfont=dict(size=14)),
                             yaxis=dict(showgrid=True,gridcolor=C["grid"],ticksuffix="%",tickfont=dict(size=14)))
            st.plotly_chart(fq, use_container_width=True)

        st.markdown('<div class="section-header">POSITIONING — YTD vs 1Y</div>', unsafe_allow_html=True)
        sdf = pipe_df.dropna(subset=["YTD","1Y"]).copy()
        sdf["S"] = sdf.apply(lambda r: display_label(r["BBG_Ticker"],r["Name"]),axis=1).astype(str)
        sec_colors = ["#3b82f6","#10b981","#f59e0b","#8b5cf6","#ef4444","#06b6d4","#f97316","#84cc16"]
        fscat = go.Figure()
        for i,(sec,grp) in enumerate(sdf.groupby("Sector")):
            fscat.add_trace(go.Scatter(x=grp["YTD"],y=grp["1Y"],mode="markers+text",name=sec,
                marker=dict(color=sec_colors[i%len(sec_colors)],size=10,opacity=0.85),
                text=grp["S"],textposition="top center",
                textfont=dict(family=C["font"],size=13,color=C["text"]),
                hovertemplate="<b>%{text}</b><br>YTD: %{x:.1f}%<br>1Y: %{y:.1f}%<extra></extra>"))
        fscat.add_hline(y=0,line_color="#334155",line_width=1)
        fscat.add_vline(x=0,line_color="#334155",line_width=1)
        fscat.update_layout(title=dict(text="Pipeline: YTD vs 1-Year",font=dict(family=C["font"],size=15,color="#64748b"),x=0),
                            **base_layout(420,margin=dict(l=10,r=10,t=44,b=60)),
                            xaxis=dict(showgrid=True,gridcolor=C["grid"],ticksuffix="%",title="YTD",tickfont=dict(size=14)),
                            yaxis=dict(showgrid=True,gridcolor=C["grid"],ticksuffix="%",title="1-Year",tickfont=dict(size=14)),
                            legend=dict(font=dict(family=C["font"],size=13,color="#64748b"),orientation="h",y=-0.2))
        st.plotly_chart(fscat, use_container_width=True)

        st.markdown('<div class="section-header">PIPELINE TABLE</div>', unsafe_allow_html=True)
        pt = pipe_df[["BBG_Ticker","Name","Sector","Quarter"]+PERIODS].copy()
        pt["Ticker"] = pt.apply(lambda r: display_label(r["BBG_Ticker"],r["Name"]),axis=1).astype(str)
        pt = pt.drop(columns=["BBG_Ticker"]).rename(columns={"Quarter":"Q"})
        pt = pt[["Ticker","Name","Sector","Q"]+PERIODS]
        st.dataframe(pt.style.applymap(style_pct,subset=PERIODS)
                     .format({p: lambda x: fmt_pct(x) for p in PERIODS})
                     .set_properties(**{"font-family":"IBM Plex Mono","font-size":"12px"}),
                     use_container_width=True, height=380)


# ═══════════════════════════════════════════════════════════════════════════════
# TAB 4 — COMPETITORS & MARKET SHARE
# ═══════════════════════════════════════════════════════════════════════════════

# ── Issuer registry ────────────────────────────────────────────────────────────
ISSUERS = {
    "KTB":       "80",
    "Bualuang":  "01",
    "Kasikorn":  "11",
    "KGI":       "13",
    "Yuanta":    "19",
    "InnovestX": "23",
    "Finansia":  "24",
    "Pi":        "03",
    "KKP":       "06",
}
ISSUER_COLORS = {
    "KTB":       "#3b82f6",
    "Bualuang":  "#f59e0b",
    "Kasikorn":  "#10b981",
    "KGI":       "#8b5cf6",
    "Yuanta":    "#ef4444",
    "InnovestX": "#06b6d4",
    "Finansia":  "#f97316",
    "Pi":        "#84cc16",
    "KKP":       "#ec4899",
}
LIQUIDITY_PERIODS = {"1W": 7, "1M": 30, "3M": 91, "6M": 182, "YTD": 0, "1Y": 365}


@st.cache_data(ttl=900, show_spinner=False)
def fetch_issuer_dr_data(base_codes: tuple, issuers: tuple, period_days: int) -> pd.DataFrame:
    """Fetch price + volume history for every base_code × issuer DR ticker on SET."""
    today    = datetime.today()
    start_dt = datetime(today.year, 1, 1) if period_days == 0 else today - timedelta(days=period_days + 5)
    start_s  = start_dt.strftime("%Y-%m-%d")
    end_s    = today.strftime("%Y-%m-%d")

    ticker_map = {f"{base}{code}.BK": (base, issuer, code)
                  for base in base_codes for issuer, code in issuers}
    if not ticker_map:
        return pd.DataFrame()

    CHUNK = 100
    hist_dict = {}  # yahoo -> DataFrame with Close/Volume
    all_tickers = list(ticker_map.keys())
    for i in range(0, len(all_tickers), CHUNK):
        chunk = all_tickers[i:i + CHUNK]
        try:
            raw = yf.download(chunk, start=start_s, end=end_s,
                              auto_adjust=True, progress=False)
            if raw.empty:
                continue
            if isinstance(raw.columns, pd.MultiIndex):
                for tkr in chunk:
                    try:
                        df_t = raw[tkr].dropna(subset=["Close"])
                        if len(df_t) > 0:
                            hist_dict[tkr] = df_t
                    except Exception:
                        continue
            else:
                # Single ticker
                df_t = raw.dropna(subset=["Close"])
                if len(df_t) > 0:
                    hist_dict[chunk[0]] = df_t
        except Exception:
            continue

    rows = []
    for yahoo, (base, issuer, code) in ticker_map.items():
        hist = hist_dict.get(yahoo)
        if hist is None or len(hist) == 0:
            continue
        for date, row in hist.iterrows():
            close = float(row["Close"]) if "Close" in row.index else None
            vol   = float(row["Volume"]) if "Volume" in row.index else 0.0
            rows.append({"Base": base, "Issuer": issuer, "IssuerCode": code,
                          "Yahoo": yahoo, "Date": pd.Timestamp(date).date(),
                          "Close": close, "Volume": vol,
                          "Turnover": (close or 0) * vol})
    return pd.DataFrame(rows)


@st.cache_data(ttl=900, show_spinner=False)
def fetch_issuer_returns(base_codes: tuple, issuers: tuple) -> pd.DataFrame:
    """Fetch WoW / MoM / YTD / 1Y returns for every base × issuer combination.
    Downloads in chunks of 100 tickers to avoid Yahoo rate limits."""
    today     = datetime.today()
    start_dt  = (today - timedelta(days=375)).strftime("%Y-%m-%d")
    end_dt    = today.strftime("%Y-%m-%d")

    ticker_map = {f"{base}{code}.BK": (base, issuer)
                  for base in base_codes for issuer, code in issuers}
    all_tickers = list(ticker_map.keys())
    if not all_tickers:
        return pd.DataFrame()

    def pct(series, days):
        s = series.dropna()
        if len(s) < 2:
            return None
        cutoff = pd.Timestamp(today.year, 1, 1) if days == 0 else pd.Timestamp(today - timedelta(days=days))
        idx = min(s.index.searchsorted(cutoff), len(s) - 1)
        sp, ep = float(s.iloc[idx]), float(s.iloc[-1])
        return round((ep / sp - 1) * 100, 2) if sp != 0 else None

    # Batch download in chunks of 100 — keeps requests manageable
    CHUNK = 100
    prices_dict = {}  # yahoo -> pd.Series
    for i in range(0, len(all_tickers), CHUNK):
        chunk = all_tickers[i:i + CHUNK]
        try:
            raw = yf.download(chunk, start=start_dt, end=end_dt,
                              auto_adjust=True, progress=False)
            if raw.empty:
                continue
            # Normalise to DataFrame of Close prices keyed by ticker
            if isinstance(raw.columns, pd.MultiIndex):
                close = raw["Close"] if "Close" in raw.columns.get_level_values(0) else pd.DataFrame()
            else:
                # Single ticker returned as flat DataFrame
                close = raw[["Close"]].rename(columns={"Close": chunk[0]}) if "Close" in raw.columns else pd.DataFrame()

            for tkr in chunk:
                if tkr in close.columns:
                    s = close[tkr].dropna()
                    if len(s) >= 2:
                        prices_dict[tkr] = s
        except Exception:
            continue

    rows = []
    for yahoo, (base, issuer) in ticker_map.items():
        s = prices_dict.get(yahoo)
        if s is None or len(s) < 2:
            continue
        rows.append({
            "Base": base, "Issuer": issuer, "Yahoo": yahoo,
            "WoW": pct(s, 7), "MoM": pct(s, 30),
            "YTD": pct(s, 0), "1Y":  pct(s, 365),
            "LastClose": float(s.iloc[-1])
        })
    return pd.DataFrame(rows)


with tab_mktshare:
    # ── Two sub-sections via sub-tabs ──────────────────────────────────────────
    sub_returns, sub_liquidity = st.tabs(["📈  Price Returns vs Competitors", "💧  Liquidity & Market Share"])

    # ── Shared controls (above sub-tabs) ──────────────────────────────────────
    if st.session_state.df is None:
        st.info("Upload your DR80 Excel file first to begin.")
        st.stop()

    dr80_securities = st.session_state.df[st.session_state.df["Is_DR80"]].copy()
    auto_bases = sorted(set(
        short_ticker(r["BBG_Ticker"]).upper().replace("80", "")
        for _, r in dr80_securities.iterrows()
        if r["BBG_Ticker"].endswith("80 TB Equity")
    ))

    st.markdown('<div class="section-header">CONTROLS</div>', unsafe_allow_html=True)
    ctrl1, ctrl2, ctrl3 = st.columns([3, 2, 2])

    with ctrl1:
        sel_bases = st.multiselect(
            "Underlyings (auto-detected from DR80 list)",
            options=auto_bases,
            default=auto_bases[:6] if len(auto_bases) >= 6 else auto_bases,
            key="mkt_bases", placeholder="Select underlyings…"
        )
        extra_raw = st.text_input(
            "➕ Add extra underlyings (comma-separated, e.g. GRAB, MELI)",
            key="mkt_extras", placeholder="GRAB, MELI, SE…"
        )
        extra_bases = [x.strip().upper() for x in extra_raw.split(",") if x.strip()]
        all_bases   = list(dict.fromkeys(sel_bases + extra_bases))

    with ctrl2:
        sel_issuers = st.multiselect(
            "Issuers to compare",
            options=list(ISSUERS.keys()),
            default=list(ISSUERS.keys()),
            key="mkt_issuers", placeholder="Select issuers…"
        )
        liq_period_fetch = "1Y"

    with ctrl3:
        st.markdown("<br>", unsafe_allow_html=True)
        fetch_live = st.button("🔄  Fetch Live Data", use_container_width=True,
                                type="primary", key="mkt_fetch")
        if st.session_state.get("mkt_last_fetch"):
            st.caption(f"Last fetched: {st.session_state['mkt_last_fetch']}")
        if not all_bases:
            st.warning("Select at least one underlying.")
        if not sel_issuers:
            st.warning("Select at least one issuer.")

    if not all_bases or not sel_issuers:
        st.stop()

    issuer_pairs = tuple((k, ISSUERS[k]) for k in sel_issuers if k in ISSUERS)
    base_tuple   = tuple(all_bases)
    period_days  = LIQUIDITY_PERIODS[liq_period_fetch]

    if fetch_live or "mkt_returns_df" not in st.session_state:
        with st.spinner(f"Fetching {len(base_tuple) * len(issuer_pairs)} tickers from Yahoo Finance…"):
            st.session_state["mkt_liq_df"]    = fetch_issuer_dr_data(base_tuple, issuer_pairs, period_days)
            st.session_state["mkt_returns_df"] = fetch_issuer_returns(base_tuple, issuer_pairs)
            st.session_state["mkt_last_fetch"] = datetime.now().strftime("%H:%M:%S")

    liq_df_raw = st.session_state.get("mkt_liq_df", pd.DataFrame())
    ret_df     = st.session_state.get("mkt_returns_df", pd.DataFrame())

    no_liq = liq_df_raw is None or len(liq_df_raw) == 0
    no_ret = ret_df is None or len(ret_df) == 0

    if no_liq and no_ret:
        st.info("No data yet — click **🔄 Fetch Live Data** above.")
        st.stop()

    # ── Global date-range slicer ────────────────────────────────────────────
    if not no_liq:
        liq_df_raw = liq_df_raw.copy()
        liq_df_raw["Date"] = pd.to_datetime(liq_df_raw["Date"])
        _min_date = liq_df_raw["Date"].min().date()
        _max_date = liq_df_raw["Date"].max().date()

        if "dr_start_val" not in st.session_state:
            st.session_state["dr_start_val"] = _max_date - timedelta(days=30)
        if "dr_end_val" not in st.session_state:
            st.session_state["dr_end_val"] = _max_date

        st.session_state["dr_start_val"] = max(_min_date, min(st.session_state["dr_start_val"], _max_date))
        st.session_state["dr_end_val"]   = max(_min_date, min(st.session_state["dr_end_val"],   _max_date))

        st.markdown('<div class="section-header">📅 DATE RANGE — LIQUIDITY ANALYSIS</div>', unsafe_allow_html=True)
        dr_col1, dr_col2, dr_col3 = st.columns([2, 2, 3])
        with dr_col1:
            dr_start = st.date_input("From",
                                     value=st.session_state["dr_start_val"],
                                     min_value=_min_date, max_value=_max_date,
                                     key="dr_start_widget")
            st.session_state["dr_start_val"] = dr_start
        with dr_col2:
            dr_end = st.date_input("To",
                                   value=st.session_state["dr_end_val"],
                                   min_value=_min_date, max_value=_max_date,
                                   key="dr_end_widget")
            st.session_state["dr_end_val"] = dr_end
        with dr_col3:
            st.markdown("<br>", unsafe_allow_html=True)
            qp_cols = st.columns(5)
            qp_labels = ["1W", "1M", "3M", "6M", "YTD"]
            qp_days   = [7, 30, 91, 182, 0]
            for _qi, (_ql, _qd) in enumerate(zip(qp_labels, qp_days)):
                if qp_cols[_qi].button(_ql, key=f"qp_{_ql}", use_container_width=True):
                    _cutoff = (datetime(datetime.today().year, 1, 1).date()
                               if _qd == 0 else (_max_date - timedelta(days=_qd)))
                    st.session_state["dr_start_val"] = _cutoff
                    st.session_state["dr_end_val"]   = _max_date
                    st.rerun()

        _period_len = max((pd.Timestamp(dr_end) - pd.Timestamp(dr_start)).days, 1)
        _prev_end   = pd.Timestamp(dr_start) - pd.Timedelta(days=1)
        _prev_start = _prev_end - pd.Timedelta(days=_period_len)

        liq_df = liq_df_raw[(liq_df_raw["Date"].dt.date >= dr_start) &
                             (liq_df_raw["Date"].dt.date <= dr_end)].copy()
        liq_df_prev = liq_df_raw[(liq_df_raw["Date"].dt.date >= _prev_start.date()) &
                                  (liq_df_raw["Date"].dt.date <= _prev_end.date())].copy()

        n_days = liq_df["Date"].nunique()
        st.caption(
            f"📅 Showing **{dr_start.strftime('%d %b %Y')} → {dr_end.strftime('%d %b %Y')}** "
            f"({n_days} trading days) · comparing vs **{_prev_start.strftime('%d %b')} → {_prev_end.strftime('%d %b %Y')}**"
        )
    else:
        liq_df = pd.DataFrame()
        liq_df_prev = pd.DataFrame()

    st.markdown("---")

    # ══════════════════════════════════════════════════════════════════════════
    # SUB-TAB A — PRICE RETURNS vs COMPETITORS
    # ══════════════════════════════════════════════════════════════════════════
    with sub_returns:
        if no_ret:
            st.info("No return data fetched yet.")
        else:
            ret_periods = ["WoW", "MoM", "YTD", "1Y"]
            ret_df2 = ret_df.copy()
            ret_df2["Label"] = ret_df2["Base"] + "  " + ret_df2["Issuer"]
            ret_df2["IsKTB"] = (ret_df2["Issuer"] == "KTB").astype(int)
            ret_df2 = ret_df2.sort_values(["Base", "IsKTB"], ascending=[True, False])

            # ── Issuer Portfolio View ──────────────────────────────────────────
            st.markdown('<div class="section-header">ISSUER PORTFOLIO VIEW</div>', unsafe_allow_html=True)
            st.caption("Select an issuer to see all their issued DRs and their equal-weighted portfolio return.")

            port_issuers = sorted(ret_df2["Issuer"].unique())
            sel_port_issuer = st.selectbox(
                "Select issuer",
                options=port_issuers,
                index=port_issuers.index("KTB") if "KTB" in port_issuers else 0,
                key="port_issuer_sel",
                format_func=lambda x: f"{x}  ({len(ret_df2[ret_df2['Issuer']==x])} DRs)"
            )

            port_df = ret_df2[ret_df2["Issuer"] == sel_port_issuer].copy()
            iss_clr = ISSUER_COLORS.get(sel_port_issuer, "#3b82f6")

            if len(port_df) == 0:
                st.info(f"No return data for {sel_port_issuer}.")
            else:
                port_avgs = {p: port_df[p].dropna().mean() for p in ret_periods}
                port_pos  = {p: int((port_df[p].dropna() >= 0).sum()) for p in ret_periods}
                port_neg  = {p: int((port_df[p].dropna() < 0).sum()) for p in ret_periods}
                port_n    = len(port_df)

                kpi_cols = st.columns(4)
                for _ki, _p in enumerate(ret_periods):
                    _avg = port_avgs[_p]
                    _vc  = C["pos"] if (pd.notna(_avg) and _avg >= 0) else C["neg"]
                    _cls = "green" if (pd.notna(_avg) and _avg >= 0) else "red"
                    kpi_cols[_ki].markdown(f"""
<div class="metric-card {_cls}">
  <div class="metric-label" style="color:{iss_clr};">{sel_port_issuer} — {_p}</div>
  <div class="metric-value" style="color:{_vc};">{fmt_pct(_avg) if pd.notna(_avg) else '—'}</div>
  <div class="metric-sub">Portfolio avg · {port_pos[_p]}↑ {port_neg[_p]}↓ of {port_n}</div>
</div>""", unsafe_allow_html=True)

                st.markdown("<br>", unsafe_allow_html=True)

                port_period_sel = st.select_slider(
                    "View period", ret_periods, value="YTD", key="port_period_sel"
                )
                port_sorted = port_df.dropna(subset=[port_period_sel]).sort_values(port_period_sel, ascending=True)

                if len(port_sorted):
                    bar_h = max(300, len(port_sorted) * 32 + 60)
                    fig_port = go.Figure()
                    fig_port.add_trace(go.Bar(
                        x=port_sorted[port_period_sel],
                        y=port_sorted["Base"].astype(str),
                        orientation="h",
                        marker_color=[C["pos"] if v >= 0 else C["neg"] for v in port_sorted[port_period_sel]],
                        marker_line_width=0,
                        text=[fmt_pct(v, 1) for v in port_sorted[port_period_sel]],
                        textposition="outside",
                        textfont=dict(family=C["font"], size=12, color=C["text"]),
                        name="Individual DR",
                        hovertemplate="<b>%{y}</b><br>%{x:.1f}%<br>" + f"{sel_port_issuer}<extra></extra>",
                    ))
                    _pavg = port_avgs[port_period_sel]
                    if pd.notna(_pavg):
                        fig_port.add_vline(
                            x=_pavg, line_color=iss_clr, line_width=2, line_dash="dash",
                            annotation_text=f"Avg {fmt_pct(_pavg,1)}",
                            annotation_font=dict(family=C["font"], size=12, color=iss_clr),
                            annotation_position="top"
                        )
                    fig_port.add_vline(x=0, line_color="#334155", line_width=1)
                    fig_port.update_layout(
                        title=dict(
                            text=f"{sel_port_issuer} Portfolio — {port_period_sel} Returns ({port_n} DRs)",
                            font=dict(family=C["font"], size=15, color="#64748b"), x=0
                        ),
                        **base_layout(bar_h, margin=dict(l=10, r=80, t=50, b=10)),
                        xaxis=dict(showgrid=True, gridcolor=C["grid"], ticksuffix="%", tickfont=dict(size=12)),
                        yaxis=dict(showgrid=False, tickfont=dict(size=13), type="category"),
                    )
                    st.plotly_chart(fig_port, use_container_width=True)

                heat_p = port_df.copy()
                heat_p = heat_p.dropna(subset=["YTD"]).sort_values("YTD", ascending=False)
                if len(heat_p):
                    z_p = heat_p[ret_periods].values.astype(float)
                    txt_p = [[fmt_pct(v, 1) if not np.isnan(v) else "—" for v in row] for row in z_p]
                    fig_ph2 = go.Figure(go.Heatmap(
                        z=z_p, x=ret_periods, y=heat_p["Base"].tolist(),
                        colorscale=[[0,"#991b1b"],[0.3,"#7f1d1d"],[0.5,"#0f172a"],[0.7,"#064e3b"],[1,"#059669"]],
                        zmid=0, text=txt_p, texttemplate="%{text}",
                        textfont=dict(family=C["font"], size=13, color="#e2e8f0"),
                        hovertemplate="<b>%{y}</b> — %{x}<br>%{z:.1f}%<extra></extra>",
                        colorbar=dict(tickfont=dict(family=C["font"],size=12,color="#94a3b8"),
                                      ticksuffix="%", thickness=12, len=0.8,
                                      title=dict(text="Return%", font=dict(family=C["font"],size=12,color="#64748b")))
                    ))
                    fig_ph2.update_layout(
                        title=dict(text=f"{sel_port_issuer} — All DR Returns Heatmap",
                                   font=dict(family=C["font"],size=15,color="#64748b"), x=0),
                        **base_layout(max(260, len(heat_p)*28+80), margin=dict(l=10,r=80,t=44,b=10)),
                        xaxis=dict(showgrid=False, tickfont=dict(size=13)),
                        yaxis=dict(showgrid=False, autorange="reversed", tickfont=dict(size=12)),
                    )
                    st.plotly_chart(fig_ph2, use_container_width=True)

                port_tbl = port_df[["Base","Yahoo","WoW","MoM","YTD","1Y","LastClose"]].copy()
                port_tbl["LastClose"] = port_tbl["LastClose"].apply(lambda v: f"{v:.2f}" if pd.notna(v) else "—")
                port_tbl.insert(0, "Issuer", sel_port_issuer)
                st.dataframe(
                    port_tbl.style
                        .applymap(style_pct, subset=["WoW","MoM","YTD","1Y"])
                        .format({c: lambda x: fmt_pct(x, 1) for c in ["WoW","MoM","YTD","1Y"]})
                        .set_properties(**{"font-family":"IBM Plex Mono","font-size":"12px"}),
                    use_container_width=True, height=400
                )

            st.markdown("---")

            # ── All-Issuers Heatmap ────────────────────────────────────────────
            st.markdown('<div class="section-header">ALL ISSUERS × ALL UNDERLYINGS — RETURN HEATMAP</div>', unsafe_allow_html=True)

            pivot = ret_df2.set_index("Label")[ret_periods].dropna(how="all")
            if len(pivot):
                z = pivot.values.astype(float)
                text_h = [[fmt_pct(v, 1) if not (v is None or np.isnan(float(v) if v is not None else float("nan"))) else "—"
                           for v in row] for row in z]
                fig_ph = go.Figure(go.Heatmap(
                    z=z, x=ret_periods, y=pivot.index.tolist(),
                    colorscale=[[0,"#991b1b"],[0.3,"#7f1d1d"],[0.5,"#0f172a"],[0.7,"#064e3b"],[1,"#059669"]],
                    zmid=0, text=text_h, texttemplate="%{text}",
                    textfont=dict(family=C["font"], size=13, color="#e2e8f0"),
                    hovertemplate="<b>%{y}</b> — %{x}<br>%{z:.1f}%<extra></extra>",
                    colorbar=dict(tickfont=dict(family=C["font"],size=12,color="#94a3b8"),
                                  ticksuffix="%", thickness=12, len=0.8,
                                  title=dict(text="Return%", font=dict(family=C["font"],size=12,color="#64748b")))
                ))
                fig_ph.update_layout(
                    title=dict(text="Return Heatmap — All Underlyings × All Issuers",
                               font=dict(family=C["font"],size=15,color="#64748b"), x=0),
                    **base_layout(max(320, len(pivot)*22+80), margin=dict(l=10,r=80,t=44,b=10)),
                    xaxis=dict(showgrid=False, tickfont=dict(size=13)),
                    yaxis=dict(showgrid=False, autorange="reversed", tickfont=dict(size=12)),
                )
                st.plotly_chart(fig_ph, use_container_width=True)

            # ── Per-underlying bar ─────────────────────────────────────────────
            st.markdown('<div class="section-header">PER-UNDERLYING BREAKDOWN</div>', unsafe_allow_html=True)
            perf_bar_period = st.select_slider("Return period", ret_periods, value="MoM", key="perf_bar_period")

            cols_per_row = 2
            bases_to_show = [b for b in all_bases if b in ret_df2["Base"].values]
            for row_start in range(0, len(bases_to_show), cols_per_row):
                cols = st.columns(cols_per_row)
                for ci, base in enumerate(bases_to_show[row_start:row_start+cols_per_row]):
                    sub = ret_df2[ret_df2["Base"]==base].dropna(subset=[perf_bar_period])
                    sub = sub.sort_values(perf_bar_period, ascending=False)
                    if not len(sub):
                        continue
                    colors = [ISSUER_COLORS.get(i, "#64748b") for i in sub["Issuer"]]
                    with cols[ci]:
                        fig_pb = go.Figure(go.Bar(
                            x=sub["Issuer"], y=sub[perf_bar_period],
                            marker_color=colors, marker_line_width=0,
                            text=[fmt_pct(v, 1) for v in sub[perf_bar_period]],
                            textposition="outside",
                            textfont=dict(family=C["font"], size=11, color=C["text"]),
                            hovertemplate="<b>%{x}</b><br>%{y:.1f}%<extra></extra>",
                        ))
                        fig_pb.add_hline(y=0, line_color="#334155", line_width=1)
                        fig_pb.update_layout(
                            title=dict(text=f"{base} — {perf_bar_period}",
                                       font=dict(family=C["font"],size=14,color="#64748b"), x=0),
                            **base_layout(240, margin=dict(l=10,r=10,t=40,b=30)),
                            xaxis=dict(showgrid=False, tickfont=dict(size=12)),
                            yaxis=dict(showgrid=True, gridcolor=C["grid"], ticksuffix="%", tickfont=dict(size=12)),
                        )
                        st.plotly_chart(fig_pb, use_container_width=True)

            # ── Full return table ──────────────────────────────────────────────
            st.markdown('<div class="section-header">FULL RETURN TABLE</div>', unsafe_allow_html=True)
            tbl = ret_df2[["Base","Issuer","Yahoo","WoW","MoM","YTD","1Y","LastClose"]].copy()
            tbl["KTB"] = tbl["Issuer"].apply(lambda x: "★" if x=="KTB" else "")
            tbl = tbl.sort_values(["Base","Issuer"]).reset_index(drop=True)
            tbl["LastClose"] = tbl["LastClose"].apply(lambda v: f"{v:.2f}" if pd.notna(v) else "—")
            st.dataframe(
                tbl.style.applymap(style_pct, subset=["WoW","MoM","YTD","1Y"])
                   .format({c: lambda x: fmt_pct(x, 1) for c in ["WoW","MoM","YTD","1Y"]})
                   .set_properties(**{"font-family":"IBM Plex Mono","font-size":"12px"}),
                use_container_width=True, height=420
            )

    with sub_liquidity:
        st.markdown('<div class="section-header">LIQUIDITY & MARKET SHARE</div>', unsafe_allow_html=True)

        if no_liq or len(liq_df) == 0:
            st.info("No liquidity data for selected date range. Adjust dates or fetch live data.")
        else:
            liq_c1, liq_c2 = st.columns([2,1])
            with liq_c1:
                liq_metric = st.radio("Metric", ["Turnover (THB)", "Volume (Units)"],
                                       horizontal=True, key="liq_metric")
            with liq_c2:
                liq_agg = st.radio("Aggregate", ["Period total", "Daily average"],
                                    horizontal=True, key="liq_agg")

            metric_col = "Turnover" if "Turnover" in liq_metric else "Volume"
            use_avg    = "average" in liq_agg.lower()
            agg_fn     = "mean" if use_avg else "sum"
            m_label    = "Avg Daily" if use_avg else f"Total (selected)"
            m_suffix   = " (THB)" if metric_col=="Turnover" else " (units)"

            def fmt_vol(v):
                if metric_col == "Turnover":
                    return f"{v/1e6:.1f}M" if v >= 1e6 else f"{v/1e3:.0f}K"
                return f"{v/1e3:.1f}K" if v >= 1e3 else str(int(v))

            # Helper: compute share table for any df slice
            def _share_tbl(df_slice, col, fn):
                if df_slice is None or len(df_slice) == 0:
                    return pd.DataFrame()
                t = (df_slice.groupby("Issuer")[col].agg(fn)
                     .reset_index().sort_values(col, ascending=False))
                total = t[col].sum()
                t["Share%"] = t[col] / total * 100 if total > 0 else 0.0
                return t

            def _share_for(tbl, issuer):
                if tbl is None or len(tbl) == 0: return 0.0
                row = tbl[tbl["Issuer"] == issuer]
                return float(row["Share%"].iloc[0]) if len(row) else 0.0

            def _pp_color(v):
                if pd.isna(v): return ""
                return f"color:{C['pos']}" if v > 0 else (f"color:{C['neg']}" if v < 0 else "")

            by_issuer   = _share_tbl(liq_df,      metric_col, agg_fn)
            by_iss_prev = _share_tbl(liq_df_prev, metric_col, agg_fn)

            liq_t1, liq_t2, liq_t3 = st.tabs(["📊 Market Share", "🔍 By Underlying", "📈 Daily Time-Series"])

            # ══════════════════════════════════════════════════════════════════
            # LIQ T1 — MARKET SHARE
            # ══════════════════════════════════════════════════════════════════
            with liq_t1:
                st.markdown('<div class="section-header">MARKET SHARE BY ISSUER</div>', unsafe_allow_html=True)

                ktb_share  = _share_for(by_issuer, "KTB")
                top_issuer = by_issuer.iloc[0]["Issuer"] if len(by_issuer) else "—"
                ktb_abs    = by_issuer[by_issuer["Issuer"]=="KTB"][metric_col].sum() if len(by_issuer) else 0
                n_active   = liq_df[liq_df[metric_col]>0]["Yahoo"].nunique()

                # ── KPI cards ─────────────────────────────────────────────────
                kc1, kc2, kc3, kc4 = st.columns(4)
                def mcard(col, label, val, sub="", color="#94a3b8"):
                    col.markdown(f"""<div class="metric-card">
                        <div class="metric-label">{label}</div>
                        <div class="metric-value" style="color:{color}">{val}</div>
                        <div class="metric-sub">{sub}</div></div>""", unsafe_allow_html=True)

                mcard(kc1, "KTB Market Share", f"{ktb_share:.1f}%",
                      f"{metric_col} · selected period", "#3b82f6")
                mcard(kc2, "Leader", top_issuer, "by volume/turnover",
                      "#3b82f6" if top_issuer=="KTB" else "#f59e0b")
                mcard(kc3, "KTB Total", fmt_vol(ktb_abs),
                      f"{'THB' if metric_col=='Turnover' else 'units'}", "#3b82f6")
                mcard(kc4, "Active Tickers", str(n_active), "with any data", "#94a3b8")

                st.markdown("<br>", unsafe_allow_html=True)

                # ── GAINER/LOSER LEADERBOARD (most prominent section) ─────────
                st.markdown('<div class="section-header">📊 SHARE GAINERS & LOSERS VS PREV PERIOD</div>', unsafe_allow_html=True)
                st.caption(f"Current period vs equal-length prior period · metric: {metric_col}")

                if len(by_iss_prev) == 0:
                    st.info("Not enough history for prior-period comparison. Try a shorter date range or fetch more data.")
                else:
                    gain_rows = []
                    for _, row in by_issuer.iterrows():
                        iss = row["Issuer"]
                        cur_share  = row["Share%"]
                        prev_share = _share_for(by_iss_prev, iss)
                        delta_pp   = cur_share - prev_share
                        cur_vol    = row[metric_col]
                        prev_vol_row = by_iss_prev[by_iss_prev["Issuer"]==iss]
                        prev_vol   = float(prev_vol_row[metric_col].iloc[0]) if len(prev_vol_row) else 0
                        delta_vol  = cur_vol - prev_vol
                        delta_vol_pct = (delta_vol / prev_vol * 100) if prev_vol > 0 else None
                        gain_rows.append({
                            "Issuer": iss,
                            "Current Share%": cur_share,
                            "Prev Share%": prev_share,
                            "Δ Share (pp)": delta_pp,
                            f"Current {metric_col}": cur_vol,
                            f"Prev {metric_col}": prev_vol,
                            f"Δ {metric_col} %": delta_vol_pct,
                        })
                    gain_df = pd.DataFrame(gain_rows).sort_values("Δ Share (pp)", ascending=False)

                    # Visual winner/loser cards
                    winner = gain_df.iloc[0]
                    loser  = gain_df.iloc[-1]
                    w_col, l_col, spark_col = st.columns([1, 1, 2])

                    w_clr = ISSUER_COLORS.get(winner["Issuer"], "#10b981")
                    l_clr = ISSUER_COLORS.get(loser["Issuer"],  "#ef4444")
                    w_col.markdown(f"""
<div class="metric-card green" style="border-left-color:{w_clr};">
  <div class="metric-label">🏆 BIGGEST GAINER</div>
  <div style="font-family:IBM Plex Mono;font-size:1.6rem;font-weight:700;color:{w_clr};">{winner['Issuer']}</div>
  <div style="font-family:IBM Plex Mono;font-size:1.2rem;color:#10b981;">+{winner['Δ Share (pp)']:.2f} pp</div>
  <div class="metric-sub">{winner['Current Share%']:.1f}% share (was {winner['Prev Share%']:.1f}%)</div>
</div>""", unsafe_allow_html=True)

                    l_col.markdown(f"""
<div class="metric-card red" style="border-left-color:{l_clr};">
  <div class="metric-label">📉 BIGGEST LOSER</div>
  <div style="font-family:IBM Plex Mono;font-size:1.6rem;font-weight:700;color:{l_clr};">{loser['Issuer']}</div>
  <div style="font-family:IBM Plex Mono;font-size:1.2rem;color:#ef4444;">{loser['Δ Share (pp)']:.2f} pp</div>
  <div class="metric-sub">{loser['Current Share%']:.1f}% share (was {loser['Prev Share%']:.1f}%)</div>
</div>""", unsafe_allow_html=True)

                    with spark_col:
                        # Waterfall-style delta bar
                        gdf_sorted = gain_df.sort_values("Δ Share (pp)", ascending=True)
                        bar_clrs   = [C["pos"] if v >= 0 else C["neg"] for v in gdf_sorted["Δ Share (pp)"]]
                        fig_gain = go.Figure(go.Bar(
                            x=gdf_sorted["Δ Share (pp)"],
                            y=gdf_sorted["Issuer"],
                            orientation="h",
                            marker_color=bar_clrs,
                            marker_line_width=0,
                            text=[f"{v:+.2f}pp" for v in gdf_sorted["Δ Share (pp)"]],
                            textposition="outside",
                            textfont=dict(family=C["font"], size=12, color=C["text"]),
                            hovertemplate="<b>%{y}</b><br>%{x:+.2f} pp<extra></extra>",
                        ))
                        fig_gain.add_vline(x=0, line_color="#334155", line_width=1)
                        fig_gain.update_layout(
                            title=dict(text="Market Share Change (pp) vs Prior Period",
                                       font=dict(family=C["font"],size=13,color="#64748b"), x=0),
                            **base_layout(max(220, len(gdf_sorted)*34+60), margin=dict(l=10,r=70,t=40,b=10)),
                            xaxis=dict(showgrid=True, gridcolor=C["grid"], ticksuffix="pp", tickfont=dict(size=11)),
                            yaxis=dict(showgrid=False, tickfont=dict(size=13), type="category"),
                        )
                        st.plotly_chart(fig_gain, use_container_width=True)

                    # Full gainer table
                    gain_disp = gain_df.copy()
                    gain_disp["Current Share%"] = gain_disp["Current Share%"].apply(lambda v: f"{v:.1f}%")
                    gain_disp["Prev Share%"]    = gain_disp["Prev Share%"].apply(lambda v: f"{v:.1f}%")
                    gain_disp[f"Δ {metric_col} %"] = gain_disp[f"Δ {metric_col} %"].apply(
                        lambda v: f"{v:+.1f}%" if pd.notna(v) else "—")
                    gain_disp[f"Current {metric_col}"] = gain_disp[f"Current {metric_col}"].apply(fmt_vol)
                    gain_disp[f"Prev {metric_col}"]    = gain_disp[f"Prev {metric_col}"].apply(fmt_vol)
                    st.dataframe(
                        gain_disp.style
                            .applymap(_pp_color, subset=["Δ Share (pp)"])
                            .format({"Δ Share (pp)": lambda v: f"{v:+.2f}pp"})
                            .set_properties(**{"font-family":"IBM Plex Mono","font-size":"12px"}),
                        use_container_width=True, hide_index=True, height=300
                    )

                st.markdown("<br>", unsafe_allow_html=True)

                # ── Donut + per-issuer change cards ───────────────────────────
                donut_col, cards_col = st.columns([1, 1])

                with donut_col:
                    donut_colors = [ISSUER_COLORS.get(i, "#64748b") for i in by_issuer["Issuer"]]
                    fig_donut = go.Figure(go.Pie(
                        labels=by_issuer["Issuer"],
                        values=by_issuer[metric_col],
                        hole=0.55,
                        marker_colors=donut_colors,
                        textinfo="label+percent",
                        textfont=dict(family=C["font"], size=12),
                        hovertemplate="<b>%{label}</b><br>%{percent}<br>%{value:,.0f}<extra></extra>",
                        pull=[0.1 if i=="KTB" else 0 for i in by_issuer["Issuer"]],
                        sort=False,
                    ))
                    fig_donut.add_annotation(
                        text=f"KTB<br><b>{ktb_share:.1f}%</b>",
                        x=0.5, y=0.5, showarrow=False,
                        font=dict(family=C["font"], size=17, color="#3b82f6"),
                    )
                    fig_donut.update_layout(
                        title=dict(text=f"Market Share — {metric_col}",
                                   font=dict(family=C["font"],size=14,color="#64748b"), x=0),
                        **base_layout(360, margin=dict(l=10,r=10,t=44,b=10)),
                        showlegend=False,
                    )
                    st.plotly_chart(fig_donut, use_container_width=True)

                with cards_col:
                    st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.7rem;text-transform:uppercase;letter-spacing:0.1em;color:#3b82f6;border-bottom:1px solid #1e2d4a;padding-bottom:6px;margin-bottom:10px;">SHARE % · CHANGE VS PRIOR PERIOD</div>', unsafe_allow_html=True)
                    issuers_ranked = by_issuer["Issuer"].tolist()
                    card_rows2 = [issuers_ranked[i:i+2] for i in range(0, len(issuers_ranked), 2)]
                    for pair in card_rows2:
                        pair_cols = st.columns(len(pair))
                        for ci2, iss in enumerate(pair):
                            cur_sh   = _share_for(by_issuer,   iss)
                            prev_sh  = _share_for(by_iss_prev, iss)
                            delta_sh = cur_sh - prev_sh
                            iss_color = ISSUER_COLORS.get(iss, "#64748b")
                            arrow = "▲" if delta_sh > 0 else ("▼" if delta_sh < 0 else "—")
                            d_clr = C["pos"] if delta_sh > 0 else (C["neg"] if delta_sh < 0 else "#64748b")
                            pair_cols[ci2].markdown(f"""
<div class="metric-card" style="padding:10px 14px;margin-bottom:6px;">
  <div style="display:flex;justify-content:space-between;align-items:center;">
    <span style="font-family:IBM Plex Mono;font-size:0.8rem;font-weight:700;color:{iss_color};">{iss}</span>
    <span style="font-family:IBM Plex Mono;font-size:1.1rem;font-weight:700;color:#e2e8f0;">{cur_sh:.1f}%</span>
  </div>
  <div style="font-family:IBM Plex Mono;font-size:0.75rem;margin-top:4px;">
    <span style="color:{d_clr};font-weight:700;">{arrow} {abs(delta_sh):.2f}pp</span>
    <span style="color:#475569;"> vs prev period ({prev_sh:.1f}%)</span>
  </div>
</div>""", unsafe_allow_html=True)

                # ── Ranked bar ─────────────────────────────────────────────────
                st.markdown("<br>", unsafe_allow_html=True)
                by_iss_asc = by_issuer.sort_values("Share%")
                colors_bar = [ISSUER_COLORS.get(i, "#64748b") for i in by_iss_asc["Issuer"]]
                fig_ms = go.Figure(go.Bar(
                    x=by_iss_asc["Share%"], y=by_iss_asc["Issuer"], orientation="h",
                    marker_color=colors_bar, marker_line_width=0,
                    text=[f"{v:.1f}%" for v in by_iss_asc["Share%"]],
                    textposition="outside",
                    textfont=dict(family=C["font"], size=12, color=C["text"]),
                    hovertemplate="<b>%{y}</b><br>%{x:.1f}%<extra></extra>",
                ))
                fig_ms.update_layout(
                    title=dict(text="Ranked Market Share %",
                               font=dict(family=C["font"],size=14,color="#64748b"), x=0),
                    **base_layout(300, margin=dict(l=10,r=60,t=44,b=10)),
                    xaxis=dict(showgrid=True, gridcolor=C["grid"], ticksuffix="%", tickfont=dict(size=12)),
                    yaxis=dict(showgrid=False, tickfont=dict(size=13), type="category"),
                )
                st.plotly_chart(fig_ms, use_container_width=True)

                # ── Issuer drill-down: top 5 underlyings + contributor analysis ─
                st.markdown('<div class="section-header">ISSUER DRILL-DOWN — TOP UNDERLYINGS & CONTRIBUTORS</div>', unsafe_allow_html=True)
                drill_issuers = by_issuer["Issuer"].tolist()
                sel_drill = st.selectbox(
                    "Select issuer",
                    options=drill_issuers,
                    index=0,
                    key="drill_issuer",
                    format_func=lambda x: f"{x}  ({_share_for(by_issuer, x):.1f}% share)"
                )

                drill_df      = liq_df[liq_df["Issuer"] == sel_drill].copy()
                drill_df_prev = liq_df_prev[liq_df_prev["Issuer"] == sel_drill].copy()
                iss_color = ISSUER_COLORS.get(sel_drill, "#64748b")

                if len(drill_df) == 0:
                    st.info(f"No data for {sel_drill} in selected period.")
                else:
                    top_tv  = (drill_df.groupby("Base")["Turnover"].agg(agg_fn)
                               .reset_index().sort_values("Turnover", ascending=False).head(5))
                    top_vol = (drill_df.groupby("Base")["Volume"].agg(agg_fn)
                               .reset_index().sort_values("Volume", ascending=False).head(5))

                    drill_c1, drill_c2 = st.columns(2)
                    with drill_c1:
                        st.markdown(f'<div style="font-family:IBM Plex Mono;font-size:0.7rem;color:{iss_color};text-transform:uppercase;margin-bottom:6px;">{sel_drill} — Top 5 by Turnover</div>', unsafe_allow_html=True)
                        fig_dt = go.Figure(go.Bar(
                            x=top_tv["Turnover"], y=top_tv["Base"], orientation="h",
                            marker_color=iss_color, marker_line_width=0,
                            text=[fmt_vol(v) for v in top_tv["Turnover"]], textposition="outside",
                            textfont=dict(family=C["font"], size=12, color=C["text"]),
                            hovertemplate="<b>%{y}</b><br>%{x:,.0f}<extra></extra>",
                        ))
                        fig_dt.update_layout(**base_layout(240, margin=dict(l=10,r=60,t=20,b=10)),
                            xaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=11)),
                            yaxis=dict(showgrid=False, tickfont=dict(size=13), autorange="reversed"))
                        st.plotly_chart(fig_dt, use_container_width=True)

                    with drill_c2:
                        st.markdown(f'<div style="font-family:IBM Plex Mono;font-size:0.7rem;color:{iss_color};text-transform:uppercase;margin-bottom:6px;">{sel_drill} — Top 5 by Volume</div>', unsafe_allow_html=True)
                        fig_dv = go.Figure(go.Bar(
                            x=top_vol["Volume"], y=top_vol["Base"], orientation="h",
                            marker_color=iss_color, marker_line_width=0,
                            text=[fmt_vol(v) for v in top_vol["Volume"]], textposition="outside",
                            textfont=dict(family=C["font"], size=12, color=C["text"]),
                            hovertemplate="<b>%{y}</b><br>%{x:,.0f}<extra></extra>",
                        ))
                        fig_dv.update_layout(**base_layout(240, margin=dict(l=10,r=60,t=20,b=10)),
                            xaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=11)),
                            yaxis=dict(showgrid=False, tickfont=dict(size=13), autorange="reversed"))
                        st.plotly_chart(fig_dv, use_container_width=True)

                    # Underlying contributor table vs prior period
                    if len(drill_df_prev) > 0:
                        st.markdown(f'<div style="font-family:IBM Plex Mono;font-size:0.7rem;color:#475569;text-transform:uppercase;margin:10px 0 6px;">UNDERLYING CONTRIBUTORS — CHANGE VS PRIOR PERIOD</div>', unsafe_allow_html=True)
                        cur_by_base  = drill_df.groupby("Base")[metric_col].agg(agg_fn).reset_index().rename(columns={metric_col: "Current"})
                        prev_by_base = drill_df_prev.groupby("Base")[metric_col].agg(agg_fn).reset_index().rename(columns={metric_col: "Previous"})
                        contrib = cur_by_base.merge(prev_by_base, on="Base", how="outer").fillna(0)
                        contrib["Δ"] = contrib["Current"] - contrib["Previous"]
                        contrib["Δ%"] = contrib.apply(
                            lambda r: (r["Δ"] / r["Previous"] * 100) if r["Previous"] > 0 else None, axis=1)
                        contrib = contrib.sort_values("Δ", ascending=False)

                        # Contribution bar
                        contrib_show = contrib.sort_values("Δ", ascending=True)
                        fig_contrib = go.Figure(go.Bar(
                            x=contrib_show["Δ"], y=contrib_show["Base"], orientation="h",
                            marker_color=[C["pos"] if v >= 0 else C["neg"] for v in contrib_show["Δ"]],
                            marker_line_width=0,
                            text=[fmt_vol(abs(v)) for v in contrib_show["Δ"]], textposition="outside",
                            textfont=dict(family=C["font"], size=11, color=C["text"]),
                            hovertemplate="<b>%{y}</b><br>Δ: %{x:,.0f}<extra></extra>",
                        ))
                        fig_contrib.add_vline(x=0, line_color="#334155", line_width=1)
                        fig_contrib.update_layout(
                            title=dict(text=f"{sel_drill} — {metric_col} change by underlying",
                                       font=dict(family=C["font"],size=13,color="#64748b"), x=0),
                            **base_layout(max(220, len(contrib_show)*32+60), margin=dict(l=10,r=70,t=40,b=10)),
                            xaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=11)),
                            yaxis=dict(showgrid=False, tickfont=dict(size=13), type="category"),
                        )
                        st.plotly_chart(fig_contrib, use_container_width=True)

                        # Table — style on raw numeric first, then format for display
                        contrib_tbl = contrib[["Base","Δ","Δ%"]].copy()
                        contrib_tbl.insert(1, "Current",  contrib["Current"].apply(fmt_vol))
                        contrib_tbl.insert(2, "Previous", contrib["Previous"].apply(fmt_vol))
                        st.dataframe(
                            contrib_tbl.style
                                .applymap(_pp_color, subset=["Δ", "Δ%"])
                                .format({"Δ": lambda v: f"{v:+,.0f}", "Δ%": lambda v: f"{v:+.1f}%" if pd.notna(v) else "—"})
                                .set_properties(**{"font-family":"IBM Plex Mono","font-size":"12px"}),
                            use_container_width=True, hide_index=True, height=300
                        )

                # Per-underlying heatmap
                st.markdown('<div class="section-header">MARKET SHARE BY UNDERLYING × ISSUER</div>', unsafe_allow_html=True)
                pivot_ms = (liq_df.groupby(["Base","Issuer"])[metric_col].agg(agg_fn).reset_index())
                base_totals = pivot_ms.groupby("Base")[metric_col].sum()
                pivot_ms["Share%"] = pivot_ms.apply(
                    lambda r: r[metric_col] / base_totals[r["Base"]] * 100 if base_totals.get(r["Base"],0) > 0 else 0,
                    axis=1)
                pivot_wide = pivot_ms.pivot(index="Issuer", columns="Base", values="Share%").fillna(0)
                issuer_order = ["KTB"] + [i for i in sel_issuers if i != "KTB"]
                pivot_wide = pivot_wide.reindex([i for i in issuer_order if i in pivot_wide.index])
                z_ms = pivot_wide.values
                text_ms = [[f"{v:.0f}%" for v in row] for row in z_ms]
                fig_ms_heat = go.Figure(go.Heatmap(
                    z=z_ms, x=pivot_wide.columns.tolist(), y=pivot_wide.index.tolist(),
                    colorscale=[[0,"#0f172a"],[0.3,"#1e3a5f"],[0.7,"#064e3b"],[1,"#059669"]],
                    zmin=0, zmax=100,
                    text=text_ms, texttemplate="%{text}",
                    textfont=dict(family=C["font"], size=13, color="#e2e8f0"),
                    hovertemplate="<b>%{y}</b> → <b>%{x}</b><br>Share: %{z:.1f}%<extra></extra>",
                    colorbar=dict(tickfont=dict(family=C["font"],size=12,color="#94a3b8"),
                                  ticksuffix="%", thickness=12,
                                  title=dict(text="Share%",font=dict(family=C["font"],size=12,color="#64748b")))
                ))
                fig_ms_heat.update_layout(
                    title=dict(text=f"Market Share % — {metric_col} per Underlying",
                               font=dict(family=C["font"],size=15,color="#64748b"), x=0),
                    **base_layout(max(280, len(pivot_wide)*30+80), margin=dict(l=10,r=80,t=44,b=10)),
                    xaxis=dict(showgrid=False, tickfont=dict(size=12)),
                    yaxis=dict(showgrid=False, tickfont=dict(size=13)),
                )
                st.plotly_chart(fig_ms_heat, use_container_width=True)

            # ══════════════════════════════════════════════════════════════════
            # LIQ T2 — BY UNDERLYING
            # ══════════════════════════════════════════════════════════════════
            with liq_t2:
                st.markdown('<div class="section-header">VOLUME / TURNOVER BY UNDERLYING</div>', unsafe_allow_html=True)
                grp = liq_df.groupby(["Base","Issuer"])[metric_col].agg(agg_fn).reset_index()

                cols_per_row = 2
                bases_list = sorted(liq_df["Base"].unique())
                for row_start in range(0, len(bases_list), cols_per_row):
                    cols = st.columns(cols_per_row)
                    for ci, base in enumerate(bases_list[row_start:row_start+cols_per_row]):
                        bsub = grp[grp["Base"]==base].sort_values(metric_col, ascending=False)
                        if not len(bsub):
                            continue
                        colors = [ISSUER_COLORS.get(i, "#64748b") for i in bsub["Issuer"]]
                        with cols[ci]:
                            fig_liq = go.Figure(go.Bar(
                                x=bsub["Issuer"], y=bsub[metric_col],
                                marker_color=colors, marker_line_width=0,
                                text=[fmt_vol(v) for v in bsub[metric_col]],
                                textposition="outside",
                                textfont=dict(family=C["font"],size=11,color=C["text"]),
                                hovertemplate="<b>%{x}</b><br>%{y:,.0f}<extra></extra>",
                            ))
                            fig_liq.update_traces(
                                marker_line_color=["#ffffff" if i=="KTB" else "rgba(0,0,0,0)" for i in bsub["Issuer"]],
                                marker_line_width=[2 if i=="KTB" else 0 for i in bsub["Issuer"]],
                            )
                            fig_liq.update_layout(
                                title=dict(text=f"{base}  —  {m_label}{m_suffix}",
                                           font=dict(family=C["font"],size=13,color="#64748b"), x=0),
                                **base_layout(230, margin=dict(l=10,r=10,t=40,b=30)),
                                xaxis=dict(showgrid=False, tickfont=dict(size=11)),
                                yaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=11)),
                            )
                            st.plotly_chart(fig_liq, use_container_width=True)

                with st.expander("📋 Full aggregated table"):
                    full_tbl = (liq_df.groupby(["Base","Issuer","Yahoo"])
                                .agg({"Volume":"sum","Turnover":"sum","Close":"last"})
                                .reset_index()
                                .sort_values(["Base","Turnover"], ascending=[True,False]))
                    full_tbl["Turnover (M THB)"] = (full_tbl["Turnover"]/1e6).round(2)
                    full_tbl["Volume (K)"]        = (full_tbl["Volume"]/1e3).round(1)
                    st.dataframe(full_tbl[["Base","Issuer","Yahoo","Close","Volume (K)","Turnover (M THB)"]],
                                 use_container_width=True, height=380)

            # ══════════════════════════════════════════════════════════════════
            # LIQ T3 — DAILY TIME-SERIES
            # ══════════════════════════════════════════════════════════════════
            with liq_t3:
                st.markdown('<div class="section-header">DAILY TIME-SERIES</div>', unsafe_allow_html=True)

                ts_base = st.selectbox("Select underlying", sorted(liq_df["Base"].unique()), key="ts_base")
                ts_sub  = liq_df[liq_df["Base"]==ts_base].copy()

                if not len(ts_sub):
                    st.info(f"No daily data for {ts_base}.")
                else:
                    fig_ts = go.Figure()
                    for issuer in sel_issuers:
                        iss_sub = ts_sub[ts_sub["Issuer"]==issuer].sort_values("Date")
                        if not len(iss_sub) or iss_sub[metric_col].sum() == 0:
                            continue
                        is_ktb = issuer == "KTB"
                        fig_ts.add_trace(go.Scatter(
                            x=iss_sub["Date"], y=iss_sub[metric_col],
                            mode="lines", name=issuer,
                            line=dict(color=ISSUER_COLORS.get(issuer,"#64748b"),
                                      width=2.5 if is_ktb else 1.5),
                            fill="tozeroy" if is_ktb else None,
                            fillcolor="rgba(59,130,246,0.06)" if is_ktb else None,
                            opacity=1.0 if is_ktb else 0.75,
                            hovertemplate=f"<b>{issuer}</b><br>%{{x}}<br>%{{y:,.0f}}<extra></extra>",
                        ))
                    fig_ts.update_layout(
                        title=dict(text=f"{ts_base} DR — Daily {metric_col}{m_suffix}",
                                   font=dict(family=C["font"],size=15,color="#64748b"), x=0),
                        **base_layout(380, margin=dict(l=10,r=10,t=44,b=50)),
                        xaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=12)),
                        yaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=12)),
                        legend=dict(font=dict(family=C["font"],size=12,color="#94a3b8"),
                                    bgcolor="rgba(0,0,0,0)", orientation="h", y=-0.22),
                        hovermode="x unified",
                    )
                    st.plotly_chart(fig_ts, use_container_width=True)

                    # Cumulative
                    st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.75rem;color:#475569;margin:10px 0 4px;">CUMULATIVE</div>', unsafe_allow_html=True)
                    fig_cum = go.Figure()
                    for issuer in sel_issuers:
                        iss_sub = ts_sub[ts_sub["Issuer"]==issuer].sort_values("Date").copy()
                        if not len(iss_sub) or iss_sub[metric_col].sum() == 0:
                            continue
                        iss_sub["Cumul"] = iss_sub[metric_col].cumsum()
                        is_ktb = issuer == "KTB"
                        fig_cum.add_trace(go.Scatter(
                            x=iss_sub["Date"], y=iss_sub["Cumul"],
                            mode="lines", name=issuer,
                            fill="tozeroy" if is_ktb else None,
                            fillcolor="rgba(59,130,246,0.08)" if is_ktb else None,
                            line=dict(color=ISSUER_COLORS.get(issuer,"#64748b"),
                                      width=2.5 if is_ktb else 1.5),
                            opacity=1.0 if is_ktb else 0.7,
                            hovertemplate=f"<b>{issuer}</b> cumulative<br>%{{x}}<br>%{{y:,.0f}}<extra></extra>",
                        ))
                    fig_cum.update_layout(
                        title=dict(text=f"{ts_base} — Cumulative {metric_col}{m_suffix}",
                                   font=dict(family=C["font"],size=14,color="#64748b"), x=0),
                        **base_layout(320, margin=dict(l=10,r=10,t=44,b=50)),
                        xaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=12)),
                        yaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=12)),
                        legend=dict(font=dict(family=C["font"],size=12,color="#94a3b8"),
                                    bgcolor="rgba(0,0,0,0)", orientation="h", y=-0.22),
                        hovermode="x unified",
                    )
                    st.plotly_chart(fig_cum, use_container_width=True)

                    # Daily market share stacked area
                    st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.75rem;color:#475569;margin:10px 0 4px;">DAILY MARKET SHARE %</div>', unsafe_allow_html=True)
                    daily_pivot = (ts_sub.groupby(["Date","Issuer"])[metric_col].sum().unstack(fill_value=0))
                    daily_totals = daily_pivot.sum(axis=1)
                    daily_share  = daily_pivot.div(daily_totals, axis=0) * 100

                    fig_stack = go.Figure()
                    issuer_order_stack = ["KTB"] + [i for i in sel_issuers if i!="KTB" and i in daily_share.columns]
                    for issuer in reversed(issuer_order_stack):
                        if issuer not in daily_share.columns:
                            continue
                        fig_stack.add_trace(go.Scatter(
                            x=daily_share.index, y=daily_share[issuer],
                            mode="lines", name=issuer,
                            stackgroup="one",
                            fillcolor=ISSUER_COLORS.get(issuer,"#64748b"),
                            line=dict(color=ISSUER_COLORS.get(issuer,"#64748b"), width=0.5),
                            hovertemplate=f"<b>{issuer}</b><br>%{{x}}<br>%{{y:.1f}}%<extra></extra>",
                        ))
                    fig_stack.update_layout(
                        title=dict(text=f"{ts_base} — Daily Market Share % (stacked)",
                                   font=dict(family=C["font"],size=14,color="#64748b"), x=0),
                        **base_layout(300, margin=dict(l=10,r=10,t=44,b=50)),
                        xaxis=dict(showgrid=True, gridcolor=C["grid"], tickfont=dict(size=12)),
                        yaxis=dict(showgrid=True, gridcolor=C["grid"], ticksuffix="%",
                                   tickfont=dict(size=12), range=[0,100]),
                        legend=dict(font=dict(family=C["font"],size=12,color="#94a3b8"),
                                    bgcolor="rgba(0,0,0,0)", orientation="h", y=-0.22),
                        hovermode="x unified",
                    )
                    st.plotly_chart(fig_stack, use_container_width=True)

# ═══════════════════════════════════════════════════════════════════════════════
# TAB 6 — ADD SECURITY
# ═══════════════════════════════════════════════════════════════════════════════
with tab_add:
    st.markdown('<div class="section-header">ADD PIPELINE SECURITY</div>', unsafe_allow_html=True)
    col_f, col_p = st.columns([1,1])

    with col_f:
        st.markdown("Bloomberg ticker is auto-converted to Yahoo Finance format for data fetching.")
        with st.form("add_form", clear_on_submit=True):
            bbg_in  = st.text_input("Bloomberg Ticker *", placeholder="e.g. AAPL US Equity, 9988 HK Equity")
            name_in = st.text_input("Company Name *", placeholder="e.g. Apple Inc")
            q_in    = st.selectbox("Target Quarter", ["Q1","Q2","Q3"])
            sec_in  = st.selectbox("Sector", SECTORS)
            fetch_on = st.checkbox("Fetch return data from Yahoo Finance", value=True)
            add_btn  = st.form_submit_button("➕ Add Security", use_container_width=True)
        if add_btn:
            if not bbg_in.strip() or not name_in.strip():
                st.error("Bloomberg ticker and company name are required.")
            elif st.session_state.df is not None and bbg_in.strip() in st.session_state.df["BBG_Ticker"].values:
                st.warning(f"⚠️ {bbg_in.strip()} already exists.")
            else:
                bbg_clean = bbg_in.strip()
                yahoo = bbg_to_yahoo(bbg_clean)
                new_row = {"BBG_Ticker":bbg_clean,"Yahoo_Ticker":yahoo,"Name":name_in.strip(),
                           "Sector":sec_in,"Quarter":q_in,"Is_DR80":False,**{p:None for p in PERIODS}}
                if fetch_on:
                    if yahoo:
                        with st.spinner(f"Fetching {yahoo}…"):
                            rets = fetch_single(yahoo)
                        new_row.update(rets)
                        st.success(f"✓ Fetched data for {yahoo}")
                    else:
                        st.warning("TB Equity tickers not available on Yahoo Finance — added without returns.")
                st.session_state.df = pd.concat([st.session_state.df, pd.DataFrame([new_row])], ignore_index=True)
                st.success(f"✓ Added **{bbg_clean}** ({name_in.strip()}) → {sec_in} / {q_in}")
                st.rerun()

    with col_p:
        st.markdown("**Ticker Conversion Preview**")
        preview_bbg = st.text_input("", placeholder="e.g. 9984 JP Equity", label_visibility="collapsed")
        if preview_bbg.strip():
            py = bbg_to_yahoo(preview_bbg.strip())
            st.markdown(f"""<div style="font-family:IBM Plex Mono;font-size:0.8rem;background:#111827;border:1px solid #1e2d4a;border-radius:8px;padding:16px;margin-bottom:16px;">
            <div style="color:#475569;font-size:0.65rem;text-transform:uppercase;margin-bottom:10px;">Conversion Result</div>
            <div style="color:#94a3b8;margin-bottom:6px;">Bloomberg: <span style="color:#e2e8f0">{preview_bbg.strip()}</span></div>
            <div style="color:#94a3b8;">Yahoo Finance: <span style="color:{'#10b981' if py else '#ef4444'}">{py or 'N/A (TB Equity)'}</span></div>
            </div>""", unsafe_allow_html=True)
        st.markdown("**Current Pipeline**")
        if st.session_state.df is not None:
            pp = st.session_state.df[~st.session_state.df["Is_DR80"]][["BBG_Ticker","Name","Sector","Quarter","YTD"]].copy()
            pp["Ticker"] = pp.apply(lambda r: display_label(r["BBG_Ticker"],r["Name"]),axis=1)
            pp["YTD"] = pp["YTD"].apply(fmt_pct)
            pp = pp.drop(columns=["BBG_Ticker"]).rename(columns={"Quarter":"Q"})
            st.dataframe(pp[["Ticker","Name","Sector","Q","YTD"]], use_container_width=True, height=360, hide_index=True)

    st.markdown("---")
    st.markdown('<div class="section-header">SAVE TO EXCEL</div>', unsafe_allow_html=True)
    st.caption("Downloads updated Excel preserving original structure with refreshed returns and new pipeline entries.")
    if st.session_state.excel_bytes and st.session_state.df is not None:
        if st.button("Generate Updated Excel"):
            with st.spinner("Writing Excel…"):
                xl = write_excel(st.session_state.excel_bytes, st.session_state.df)
            st.download_button("⬇ Download Updated DR80_Tracking.xlsx", data=xl,
                               file_name=f"DR80_Tracking_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                               mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    else:
        st.info("Load a file in the sidebar first.")


# ══════════════════════════════════════════════════════════════════════════════
# DR ISSUERS & UNDERLYINGS CONFIG
# ══════════════════════════════════════════════════════════════════════════════
DR_ISSUERS = {
    "01": "Bualuang",
    "03": "Pi",
    "06": "KKP",
    "11": "KBank",
    "13": "KGI",
    "19": "Yuanta",
    "23": "InnovestX",
    "24": "Finansia",
    "80": "KTB",
}

DR_UNDERLYINGS = [
    "NVDA","AAPL","TSLA","META","MSFT","AMZN","GOOG","GOOGL","NFLX",
    "AMD","AVGO","QCOM","ORCL","CRM","NOW","ADBE","PLTR","CRWD","PANW",
    "MA","V","PYPL","BKNG","ABNB",
    "JNJ","LLY","AMGN","ABBV","ISRG",
    "SONY","NINTENDO","TOYOTA","SOFTBANK",
    "BABA","JD","TENCENT","MEITUAN","XIAOMI",
    "WMT","COSTCO","NKE","SBUX","KO","PEP",
    "GS","MS","JPM","BAC","BLK",
    "GOLD","NEM",
    "DBS","UOB","GRAB",
    "DELL","IBM","CSCO","MU","MRVL",
]

# Candidates scanned on every refresh to detect newly listed DRs
_DISCOVERY_EXTRA = [
    "INTC","TXN","SNOW","DDOG","ZS","WDAY","TEAM","SQ","AFRM","SPOT",
    "RBLX","RIVN","NIO","LI","XPEV","F","GM","BA","LMT","RTX","NOC",
    "CVX","XOM","COP","SLB","BHP","RIO","FCX","PFE","MRK","BMY","GILD",
    "HD","LOW","TGT","DIS","CMCSA","T","VZ","AMT","PLD","SPG","NEE",
    "SPY","QQQ","GLD","SLV","VNM","MCHI","EWT","EWJ","TIGR","FUTU",
    "GRAB","SEA","HDB","TCS","INFY","VALE","PBR",
]

BENCH_PERIODS = {
    "1D":"1d","5D":"5d","1M":"1mo","3M":"3mo",
    "6M":"6mo","1Y":"1y","3Y":"3y","5Y":"5y"
}

# ══════════════════════════════════════════════════════════════════════════════
# SUPABASE HELPERS
# ══════════════════════════════════════════════════════════════════════════════
_SUPABASE_URL  = st.secrets.get("SUPABASE_URL", "")
_ANTHROPIC_KEY = st.secrets.get("ANTHROPIC_API_KEY", "")


@st.cache_resource
def _get_conn():
    if not _SUPABASE_URL:
        return None
    try:
        conn = psycopg2.connect(_SUPABASE_URL, connect_timeout=10)
        conn.autocommit = True
        return conn
    except Exception:
        return None


def _cur():
    conn = _get_conn()
    if conn is None:
        return None
    try:
        if conn.closed:
            _get_conn.clear()
            conn = _get_conn()
        return conn.cursor(cursor_factory=RealDictCursor)
    except Exception:
        return None


def db_ensure_table():
    c = _cur()
    if c is None:
        return
    try:
        c.execute("""
            CREATE TABLE IF NOT EXISTS dr_daily (
                date        TEXT PRIMARY KEY,
                week_label  TEXT,
                total_dr    INTEGER,
                ktb_dr      INTEGER,
                set_vol     FLOAT,
                set_val     FLOAT,
                dr_vol      FLOAT,
                dr_val      FLOAT,
                ktb_vol     FLOAT,
                ktb_val     FLOAT,
                source      TEXT,
                captured_at TEXT
            )
        """)
    except Exception:
        pass


@st.cache_data(ttl=60)
def db_load() -> pd.DataFrame:
    c = _cur()
    if c is None:
        return pd.DataFrame()
    try:
        c.execute("SELECT * FROM dr_daily ORDER BY date DESC")
        rows = c.fetchall()
        return pd.DataFrame([dict(r) for r in rows]) if rows else pd.DataFrame()
    except Exception:
        return pd.DataFrame()


def db_upsert(row: dict) -> bool:
    c = _cur()
    if c is None:
        st.error("Database not connected — check SUPABASE_URL secret.")
        return False
    try:
        c.execute("""
            INSERT INTO dr_daily
                (date,week_label,total_dr,ktb_dr,set_vol,set_val,
                 dr_vol,dr_val,ktb_vol,ktb_val,source,captured_at)
            VALUES
                (%(date)s,%(week_label)s,%(total_dr)s,%(ktb_dr)s,%(set_vol)s,%(set_val)s,
                 %(dr_vol)s,%(dr_val)s,%(ktb_vol)s,%(ktb_val)s,%(source)s,%(captured_at)s)
            ON CONFLICT (date) DO UPDATE SET
                week_label  = EXCLUDED.week_label,
                total_dr    = EXCLUDED.total_dr,
                ktb_dr      = EXCLUDED.ktb_dr,
                set_vol     = EXCLUDED.set_vol,
                set_val     = EXCLUDED.set_val,
                dr_vol      = EXCLUDED.dr_vol,
                dr_val      = EXCLUDED.dr_val,
                ktb_vol     = EXCLUDED.ktb_vol,
                ktb_val     = EXCLUDED.ktb_val,
                source      = EXCLUDED.source,
                captured_at = EXCLUDED.captured_at
        """, row)
        db_load.clear()
        return True
    except Exception as e:
        st.error(f"Save failed: {e}")
        return False


def db_delete(date_str: str):
    c = _cur()
    if c is None:
        return
    try:
        c.execute("DELETE FROM dr_daily WHERE date = %s", (date_str,))
        db_load.clear()
    except Exception:
        pass


def _week_label(dt: datetime) -> str:
    mon = dt - timedelta(days=dt.weekday())
    sun = mon + timedelta(days=6)
    return f"{mon.strftime('%d %b')}–{sun.strftime('%d %b %Y')}"


# ══════════════════════════════════════════════════════════════════════════════
# AI SCREENSHOT EXTRACTION
# ══════════════════════════════════════════════════════════════════════════════
def extract_from_screenshot(img_bytes: bytes) -> dict:
    if not _ANTHROPIC_KEY:
        return {}
    b64 = base64.standard_b64encode(img_bytes).decode()
    payload = {
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 512,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image",
                 "source": {"type": "base64", "media_type": "image/png", "data": b64}},
                {"type": "text", "text": (
                    "This is a screenshot of the SET Thailand DR market data page. "
                    "Extract and return ONLY a JSON object with keys: "
                    "total_dr (int), ktb_dr (int), dr_vol (float), dr_val (float), "
                    "ktb_vol (float), ktb_val (float), "
                    "set_vol (float, 0 if not shown), set_val (float, 0 if not shown). "
                    "Return ONLY valid JSON, no explanation or markdown."
                )}
            ]
        }]
    }
    try:
        r = requests.post(
            "https://api.anthropic.com/v1/messages",
            headers={
                "x-api-key": _ANTHROPIC_KEY,
                "anthropic-version": "2023-06-01",
                "content-type": "application/json"
            },
            json=payload, timeout=30
        )
        text = r.json()["content"][0]["text"].strip()
        text = re.sub(r"```json|```", "", text).strip()
        return json.loads(text)
    except Exception:
        return {}


# ══════════════════════════════════════════════════════════════════════════════
# YAHOO FINANCE — DR PRICE & BENCHMARKING FETCHERS
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(ttl=300)
def fetch_all_dr_prices(underlyings: tuple, issuers: tuple) -> pd.DataFrame:
    tickers = [f"{u}{c}.BK" for u in underlyings for c in issuers]
    try:
        raw = yf.download(tickers, period="2d", interval="1d",
                          auto_adjust=True, progress=False)
        if raw.empty:
            return pd.DataFrame()
        close = raw["Close"] if "Close" in raw.columns else raw.get("close", pd.DataFrame())
        vol   = raw["Volume"] if "Volume" in raw.columns else raw.get("volume", pd.DataFrame())
        rows = []
        for sym in tickers:
            try:
                sc = close[sym].dropna() if sym in close.columns else pd.Series(dtype=float)
                sv = vol[sym].dropna()   if sym in vol.columns   else pd.Series(dtype=float)
                if len(sc) == 0:
                    continue
                price  = float(sc.iloc[-1])
                prev   = float(sc.iloc[-2]) if len(sc) > 1 else price
                volume = int(sv.iloc[-1])   if len(sv) > 0 else 0
                chg    = (price - prev) / prev * 100 if prev else 0
                code   = sym.replace(".BK", "")[-2:]
                rows.append({
                    "Symbol":     sym,
                    "Underlying": sym.replace(".BK", "")[:-2],
                    "Issuer":     DR_ISSUERS.get(code, code),
                    "Code":       code,
                    "Price":      round(price, 4),
                    "Chg %":      round(chg, 2),
                    "Volume":     volume,
                    "Value_proxy": round(price * volume / 1000, 2),
                })
            except Exception:
                continue
        return pd.DataFrame(rows)
    except Exception as e:
        st.error(f"Yahoo Finance error: {e}")
        return pd.DataFrame()


@st.cache_data(ttl=600)
def fetch_period_returns(underlyings: tuple, issuers: tuple, period: str) -> pd.DataFrame:
    tickers = [f"{u}{c}.BK" for u in underlyings for c in issuers]
    try:
        raw = yf.download(tickers, period=period, interval="1d",
                          auto_adjust=True, progress=False)
        if raw.empty:
            return pd.DataFrame()
        close = raw["Close"] if "Close" in raw.columns else raw.get("close", pd.DataFrame())
        rows = []
        for sym in tickers:
            try:
                s = close[sym].dropna() if sym in close.columns else pd.Series(dtype=float)
                if len(s) < 2:
                    continue
                ret  = (s.iloc[-1] - s.iloc[0]) / s.iloc[0] * 100
                code = sym.replace(".BK", "")[-2:]
                rows.append({
                    "Underlying": sym.replace(".BK", "")[:-2],
                    "Issuer":     DR_ISSUERS.get(code, code),
                    "Return %":   round(float(ret), 2),
                })
            except Exception:
                continue
        return pd.DataFrame(rows)
    except Exception:
        return pd.DataFrame()


@st.cache_data(ttl=3600)
def discover_new_drs(known: tuple, issuers: tuple) -> list:
    """Scan extra candidates not in known list — alert if any exist on Yahoo."""
    candidates = [u for u in _DISCOVERY_EXTRA if u not in known]
    if not candidates:
        return []
    tickers = [f"{u}{c}.BK" for u in candidates for c in issuers]
    try:
        raw = yf.download(tickers, period="5d", interval="1d",
                          auto_adjust=True, progress=False)
        if raw.empty:
            return []
        close = raw["Close"] if "Close" in raw.columns else raw.get("close", pd.DataFrame())
        return [sym for sym in tickers
                if sym in close.columns and len(close[sym].dropna()) > 0]
    except Exception:
        return []


@st.cache_data(ttl=600)
def fetch_wow(underlyings: tuple, issuers: tuple) -> pd.DataFrame:
    tickers = [f"{u}{c}.BK" for u in underlyings for c in issuers]
    try:
        raw = yf.download(tickers, period="14d", interval="1d",
                          auto_adjust=True, progress=False)
        if raw.empty:
            return pd.DataFrame()
        close = raw["Close"] if "Close" in raw.columns else raw.get("close", pd.DataFrame())
        vol   = raw["Volume"] if "Volume" in raw.columns else raw.get("volume", pd.DataFrame())
        rows = []
        for sym in tickers:
            try:
                sc = close[sym].dropna() if sym in close.columns else pd.Series(dtype=float)
                sv = vol[sym].dropna()   if sym in vol.columns   else pd.Series(dtype=float)
                if len(sc) < 6:
                    continue
                val       = sc * sv / 1000
                this_week = float(val.iloc[-5:].sum())
                last_week = float(val.iloc[-10:-5].sum())
                wow       = (this_week - last_week) / last_week * 100 if last_week else 0
                code      = sym.replace(".BK", "")[-2:]
                rows.append({
                    "Issuer":            DR_ISSUERS.get(code, code),
                    "Underlying":        sym.replace(".BK", "")[:-2],
                    "This Week ('000)":  round(this_week, 0),
                    "Last Week ('000)":  round(last_week, 0),
                    "WoW %":             round(wow, 2),
                })
            except Exception:
                continue
        return pd.DataFrame(rows)
    except Exception:
        return pd.DataFrame()


# ══════════════════════════════════════════════════════════════════════════════
# TAB 6 — KTB DR MARKET SHARE TRACKER
# ══════════════════════════════════════════════════════════════════════════════
# ── Footer ─────────────────────────────────────────────────────────────────────
st.markdown("---")
st.markdown('<div style="font-family:IBM Plex Mono;font-size:0.6rem;color:#1e2d4a;text-align:center;">KTB SECURITIES · DR OPERATIONS · DR80 TRACKING SYSTEM</div>', unsafe_allow_html=True)
